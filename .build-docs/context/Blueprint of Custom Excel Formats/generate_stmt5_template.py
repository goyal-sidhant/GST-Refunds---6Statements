#!/usr/bin/env python3
"""
GST Refund Offline Utility — Template Generator
Statement 5 — SEZ without Payment of Tax (SEZWOP)

Generates: GST_Refund_Stmt5_Template.xlsx
Usage:     python generate_stmt5_template.py
Requires:  pip install openpyxl

The generated template follows OctaSales design language:
  - Blue headers (0070C0) = Mandatory
  - Green headers (00B050) = Optional  
  - Light blue headers (00B0F0) = Conditional
  - Calibri font throughout
  - Data starts at row 2 (no gap row)
  - Dropdowns on type/classification columns
"""
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

wb = openpyxl.Workbook()

FILL_REQ = PatternFill('solid', fgColor='0070C0')
FILL_OPT = PatternFill('solid', fgColor='00B050')
FILL_COND = PatternFill('solid', fgColor='00B0F0')
FILL_SECTION = PatternFill('solid', fgColor='FFC000')
FILL_TITLE = PatternFill('solid', fgColor='0070C0')

FW = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
FW11 = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
FB11 = Font(name='Calibri', bold=True, size=11)
FN11 = Font(name='Calibri', size=11)
FN10 = Font(name='Calibri', size=10)
FB10 = Font(name='Calibri', bold=True, size=10)
FTITLE = Font(name='Calibri', bold=True, color='FFFFFF', size=20)
FSEC = Font(name='Calibri', bold=True, size=14)
FHELP_H = Font(name='Calibri', bold=True, size=11)
FMONO = Font(name='Consolas', size=10)

AC = Alignment(horizontal='center', vertical='center', wrap_text=True)
ALW = Alignment(horizontal='left', vertical='top', wrap_text=True)
TB = Border(left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9'))

def help_row(ws, row, cat, name, ftype, desc, val=""):
    fills = {"Required": FILL_REQ, "Optional": FILL_OPT, "Conditional": FILL_COND}
    for ci, v in enumerate([cat, name, ftype, desc, val]):
        cell = ws.cell(row=row, column=ci+1, value=v)
        cell.font = FN10; cell.border = TB; cell.alignment = ALW
        if ci == 0 and cat in fills:
            cell.fill = fills[cat]
            cell.font = Font(name='Calibri', size=10, color='FFFFFF')
    return row + 1

def help_section(ws, r, title):
    ws.merge_cells(f'A{r}:E{r}')
    ws[f'A{r}'] = title; ws[f'A{r}'].font = FSEC; ws[f'A{r}'].fill = FILL_SECTION
    return r + 1

def help_desc(ws, r, text):
    ws.merge_cells(f'A{r}:E{r}')
    ws[f'A{r}'] = text; ws[f'A{r}'].font = FN11
    return r + 1

def help_table_header(ws, r):
    for ci, h in enumerate(["Field Category", "Field Name", "Field Type", "Description", "Validation"]):
        ws.cell(row=r, column=ci+1, value=h).font = FHELP_H
        ws.cell(row=r, column=ci+1).border = TB
    return r + 1

def help_note(ws, r, text):
    ws.merge_cells(f'A{r}:E{r}')
    ws[f'A{r}'] = text; ws[f'A{r}'].font = FN10
    return r + 1

# ═══════════════════════════════════════
#  Overview
# ═══════════════════════════════════════
ws = wb.active
ws.title = "Overview"
ws.sheet_properties.tabColor = "0070C0"
ws.column_dimensions['A'].width = 3
ws.column_dimensions['B'].width = 25
ws.column_dimensions['C'].width = 65

ws.merge_cells('B2:C2')
ws['B2'] = "Refund Statement 5 Template"
ws['B2'].font = FTITLE; ws['B2'].fill = FILL_TITLE

for r, (label, val) in enumerate([
    ("Refund Type", "Supplies made to SEZ Unit/SEZ Developer — without Payment of Tax"),
    ("Legal Basis", "Section 54(3) CGST Act, Rule 89(2)(d), Rule 89(2)(e)"),
    ("Refund Code / Version", "SEZWOP / 2.0"),
], start=5):
    ws[f'B{r}'] = label; ws[f'C{r}'] = val
    ws[f'B{r}'].font = FW11; ws[f'B{r}'].fill = FILL_REQ; ws[f'C{r}'].font = FN11

ws['B9'] = "Sheet Name"; ws['C9'] = "Description"
ws['B9'].font = FW11; ws['B9'].fill = FILL_REQ
ws['C9'].font = FW11; ws['C9'].fill = FILL_REQ

for i, (s, d) in enumerate([
    ("Header", "GSTIN and Return Period details"),
    ("Goods", "Supply of goods to SEZ — invoices and shipping/endorsed invoice details"),
    ("Services", "Supply of services to SEZ — invoices and shipping/endorsed invoice details"),
    ("Help", "Field specifications, validation rules, and usage guidance"),
]):
    ws[f'B{10+i}'] = s; ws[f'C{10+i}'] = d
    ws[f'B{10+i}'].font = FW11; ws[f'B{10+i}'].fill = FILL_REQ; ws[f'C{10+i}'].font = FN11

ws['B15'] = "Colour Coding"; ws['B15'].font = FB11
ws.merge_cells('B16:C16'); ws['B16'] = "Mandatory Columns"; ws['B16'].font = FW11; ws['B16'].fill = FILL_REQ
ws.merge_cells('B17:C17'); ws['B17'] = "Optional Columns"; ws['B17'].font = FW11; ws['B17'].fill = FILL_OPT
ws.merge_cells('B18:C18'); ws['B18'] = "Conditional Columns"; ws['B18'].font = FW11; ws['B18'].fill = FILL_COND

ws.merge_cells('B20:C20')
ws['B20'] = "Refund template for Supplies to SEZ Unit/Developer without Payment of Tax"
ws['B20'].font = FB11

# ═══════════════════════════════════════
#  Header
# ═══════════════════════════════════════
ws_hd = wb.create_sheet("Header")
ws_hd.sheet_properties.tabColor = "0070C0"
ws_hd.column_dimensions['A'].width = 28
ws_hd.column_dimensions['B'].width = 30
ws_hd.column_dimensions['C'].width = 50

for ci, v in enumerate(["Field Name", "Value", "Description"]):
    c = ws_hd.cell(row=1, column=ci+1, value=v)
    c.font = FW; c.fill = FILL_REQ; c.alignment = AC; c.border = TB

fields = [
    ("GSTIN", "", "15-character GSTIN of the refund applicant"),
    ("From Period (MM-YYYY)", "", "Starting return period e.g., 07-2024 for July 2024"),
    ("To Period (MM-YYYY)", "", "Ending return period e.g., 09-2024 for September 2024"),
]
for i, (field, val, desc) in enumerate(fields):
    r = i + 2
    ws_hd.cell(row=r, column=1, value=field).font = FB10
    ws_hd.cell(row=r, column=2, value=val).font = FN10
    ws_hd.cell(row=r, column=3, value=desc).font = FN10
    for c in range(1, 4): ws_hd.cell(row=r, column=c).border = TB

# ═══════════════════════════════════════
#  Goods
# ═══════════════════════════════════════
ws_g = wb.create_sheet("Goods")
ws_g.sheet_properties.tabColor = "27AE60"

goods_cols = [
    ("Doc Type", 14, FILL_REQ),
    ("Doc No", 18, FILL_REQ),
    ("Doc Date", 14, FILL_REQ),
    ("Doc Value", 16, FILL_REQ),
    ("SB / Endorsed Invoice No", 24, FILL_OPT),
    ("SB / Endorsed Invoice Date", 24, FILL_COND),
]

for ci, (name, w, fill) in enumerate(goods_cols):
    col = ci + 1
    ws_g.column_dimensions[get_column_letter(col)].width = w
    cell = ws_g.cell(row=1, column=col, value=name)
    cell.font = FW; cell.fill = fill; cell.alignment = AC; cell.border = TB

# TEXT format on text-type columns — preserves leading zeros
for text_col in ['B', 'E']:  # Doc No, SB/Endorsed Invoice No
    ws_g.column_dimensions[text_col].number_format = '@'

sample_g = ["Invoice", "SEZ/G/001", "10-09-2024", 500000, "SB2024001", "12-09-2024"]
for ci, v in enumerate(sample_g):
    cell = ws_g.cell(row=2, column=ci+1, value=v)
    cell.font = FN10; cell.border = TB

dv_doc_g = DataValidation(type="list", formula1='"Invoice,Debit Note,Credit Note"', allow_blank=True)
ws_g.add_data_validation(dv_doc_g)
dv_doc_g.add('A2:A10001')
ws_g.freeze_panes = 'A2'

# ═══════════════════════════════════════
#  Services
# ═══════════════════════════════════════
ws_s = wb.create_sheet("Services")
ws_s.sheet_properties.tabColor = "E67E22"

svc_cols = [
    ("Doc Type", 14, FILL_REQ),
    ("Doc No", 18, FILL_REQ),
    ("Doc Date", 14, FILL_REQ),
    ("Doc Value", 16, FILL_REQ),
    ("SB / Endorsed Invoice No", 24, FILL_OPT),
    ("SB / Endorsed Invoice Date", 24, FILL_COND),
]

for ci, (name, w, fill) in enumerate(svc_cols):
    col = ci + 1
    ws_s.column_dimensions[get_column_letter(col)].width = w
    cell = ws_s.cell(row=1, column=col, value=name)
    cell.font = FW; cell.fill = fill; cell.alignment = AC; cell.border = TB

# TEXT format on text-type columns — preserves leading zeros
for text_col in ['B', 'E']:  # Doc No, SB/Endorsed Invoice No
    ws_s.column_dimensions[text_col].number_format = '@'

sample_s = ["Invoice", "SEZ/S/001", "15-09-2024", 300000, "", ""]
for ci, v in enumerate(sample_s):
    cell = ws_s.cell(row=2, column=ci+1, value=v if v != "" else None)
    cell.font = FN10; cell.border = TB

dv_doc_s = DataValidation(type="list", formula1='"Invoice,Debit Note,Credit Note"', allow_blank=True)
ws_s.add_data_validation(dv_doc_s)
dv_doc_s.add('A2:A10001')
ws_s.freeze_panes = 'A2'

# ═══════════════════════════════════════
#  Help
# ═══════════════════════════════════════
ws_h = wb.create_sheet("Help")
ws_h.sheet_properties.tabColor = "FFC000"
ws_h.column_dimensions['A'].width = 14
ws_h.column_dimensions['B'].width = 24
ws_h.column_dimensions['C'].width = 18
ws_h.column_dimensions['D'].width = 75
ws_h.column_dimensions['E'].width = 25

r = 2

# Section 1: Header
r = help_section(ws_h, r, "1. Sheet Name: Header")
r = help_desc(ws_h, r, "The following table specifies the header fields for the refund application.")
r = help_table_header(ws_h, r)
r = help_row(ws_h, r, "Required", "GSTIN", "Text (15)", "15-character GSTIN of the refund applicant.", "GSTIN regex; auto upper-cased and trimmed")
r = help_row(ws_h, r, "Required", "From Period", "Text (MM-YYYY)", "Starting return period.\nExample: 07-2024 for July 2024", "Must be >= 07-2017; <= current; <= To Period")
r = help_row(ws_h, r, "Required", "To Period", "Text (MM-YYYY)", "Ending return period.\nExample: 09-2024 for September 2024", "Must be >= 07-2017; <= current; >= From Period")
r += 1

# Section 2: Goods
r = help_section(ws_h, r, "2. Sheet Name: Goods")
r = help_desc(ws_h, r, "Details of goods supplied to SEZ without payment of tax. One row = one supply document. The G/S flag is implicit — this sheet = Goods.")
r = help_table_header(ws_h, r)
r = help_row(ws_h, r, "Required", "Doc Type", "Text", "Type of document.\nValid values: Invoice, Debit Note, Credit Note", "Dropdown provided")
r = help_row(ws_h, r, "Required", "Doc No", "Text (16)", "Document number.\nMax 16 alphanumeric characters, / and - allowed.", "Cannot be 0 or special chars only")
r = help_row(ws_h, r, "Required", "Doc Date", "Text (DD-MM-YYYY)", "Date of document.\nMust be on or after 01-07-2017. Cannot be future.", "Valid calendar date; <= To Period")
r = help_row(ws_h, r, "Required", "Doc Value", "Number (15,2)", "Total document value.\nMust be >= 0.", "Non-negative")
r = help_row(ws_h, r, "Optional", "SB / Endorsed Invoice No", "Text", "Shipping Bill or Endorsed Invoice by SEZ number.\nOptional — provide if available.", "Alphanumeric if provided")
r = help_row(ws_h, r, "Conditional", "SB / Endorsed Invoice Date", "Text (DD-MM-YYYY)", "Date of Shipping Bill or Endorsed Invoice.\nRequired if SB No is provided. Cannot be future.", "Valid calendar date; required if No is filled")
r += 1

# Section 3: Services
r = help_section(ws_h, r, "3. Sheet Name: Services")
r = help_desc(ws_h, r, "Details of services supplied to SEZ without payment of tax. One row = one supply document. The G/S flag is implicit — this sheet = Services.")
r = help_table_header(ws_h, r)
r = help_row(ws_h, r, "Required", "Doc Type", "Text", "Type of document.\nValid values: Invoice, Debit Note, Credit Note", "Dropdown provided")
r = help_row(ws_h, r, "Required", "Doc No", "Text (16)", "Document number.\nMax 16 alphanumeric characters, / and - allowed.", "Cannot be 0 or special chars only")
r = help_row(ws_h, r, "Required", "Doc Date", "Text (DD-MM-YYYY)", "Date of document.\nMust be on or after 01-07-2017. Cannot be future.", "Valid calendar date; <= To Period")
r = help_row(ws_h, r, "Required", "Doc Value", "Number (15,2)", "Total document value.\nMust be >= 0.", "Non-negative")
r = help_row(ws_h, r, "Optional", "SB / Endorsed Invoice No", "Text", "Shipping Bill or Endorsed Invoice by SEZ number.\nOptional for services.", "Alphanumeric if provided")
r = help_row(ws_h, r, "Conditional", "SB / Endorsed Invoice Date", "Text (DD-MM-YYYY)", "Date of Shipping Bill or Endorsed Invoice.\nRequired if SB No is provided. Cannot be future.", "Valid calendar date; required if No is filled")
r += 1

# Section 4: Key differences
r = help_section(ws_h, r, "4. Key Differences from Other Statements")
r += 1
diffs = [
    "• This is the LIGHTEST template — only 6 columns per sheet. No tax amount columns.",
    "• Refund category: Accumulated ITC (no output tax was paid on the SEZ supply).",
    "• No BRC/FIRC block — unlike S03 (exports without payment), S05 has no banking evidence fields.",
    "• No Port Code, FOB Value, EGM fields — unlike S03 (export goods).",
    "• Goods and Services sheets have IDENTICAL columns — the only difference is the G/S type flag in JSON.",
    "• SB / Endorsed Invoice is optional for both Goods and Services (unlike S03 where Port Code/SB is mandatory for goods).",
]
for d in diffs:
    r = help_note(ws_h, r, d)
r += 1

# Section 5: Validation Rules
r = help_section(ws_h, r, "5. Key Validation Rules")
r += 1
rules = [
    "• Doc Date must be >= 01-07-2017 and <= current date and <= To Period.",
    "• SB Date cannot be future. No lower bound on SB Date.",
    "• SB Date is required only when SB No is provided.",
    "• Duplicate key = Doc Type suffix (last 4 chars) + Doc No + Financial Year.",
    "• If invoice month is Jan-Mar, FY year is decremented by 1.",
    "• App auto-generates serial numbers across both sheets (Goods first, then Services).",
    "• Return Periods (From/To) ARE required for S05 (v2.0 — unlike S02/S04 which are v3.0).",
]
for rule in rules:
    r = help_note(ws_h, r, rule)
r += 1

# Section 6: Running the Tool
r = help_section(ws_h, r, "6. Running the Tool")
r += 1
for cmd in [
    "python main.py                         → Opens file picker",
    "python main.py input.xlsx              → Process specific file",
    "python main.py input.xlsx --pretty     → Pretty-print JSON output",
    "python main.py input.xlsx -v           → Verbose validation output",
]:
    ws_h.merge_cells(f'A{r}:E{r}')
    ws_h[f'A{r}'] = f"  {cmd}"; ws_h[f'A{r}'].font = FMONO; r += 1

ws_h.freeze_panes = 'A3'
wb.active = wb.sheetnames.index("Overview")
wb.save(os.path.join(os.getcwd(), "GST_Refund_Stmt5_Template.xlsx"))
output = os.path.join(os.getcwd(), "GST_Refund_Stmt5_Template.xlsx")
print(f"Created: {output}")
