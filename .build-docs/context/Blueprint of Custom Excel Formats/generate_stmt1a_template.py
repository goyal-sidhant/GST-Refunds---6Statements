#!/usr/bin/env python3
"""
GST Refund Offline Utility — Template Generator
Statement 1A — Inverted Tax Structure (INVITC)

Generates: GST_Refund_Stmt1A_Template.xlsx
Usage:     python generate_stmt1a_template.py
Requires:  pip install openpyxl

The generated template follows OctaSales design language:
  - Blue headers (0070C0) = Mandatory
  - Green headers (00B050) = Optional  
  - Light blue headers (00B0F0) = Conditional
  - Calibri font throughout
  - Data starts at row 2 (no gap row)
  - Dropdowns on type/classification columns
"""
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from octasales_refund_style import apply_octa_replica_styling

wb = openpyxl.Workbook()

# ─── OctaSales exact colour scheme ───
FILL_REQUIRED = PatternFill('solid', fgColor='0070C0')       # Blue — mandatory
FILL_OPTIONAL = PatternFill('solid', fgColor='00B050')       # Green — optional
FILL_CONDITIONAL = PatternFill('solid', fgColor='00B0F0')    # Light blue — conditional
FILL_SECTION_HEADER = PatternFill('solid', fgColor='FFC000') # Amber — Help section titles
FILL_TITLE = PatternFill('solid', fgColor='0070C0')          # Blue — Overview title
FILL_NONE = PatternFill(fill_type=None)

FONT_WHITE_BOLD = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
FONT_WHITE_BOLD_11 = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
FONT_BOLD_11 = Font(name='Calibri', bold=True, size=11)
FONT_BOLD_10 = Font(name='Calibri', bold=True, size=10)
FONT_NORMAL_11 = Font(name='Calibri', size=11)
FONT_NORMAL_10 = Font(name='Calibri', size=10)
FONT_TITLE = Font(name='Calibri', bold=True, color='FFFFFF', size=20)
FONT_SECTION = Font(name='Calibri', bold=True, size=14)
FONT_HELP_TABLE_HEADER = Font(name='Calibri', bold=True, size=11)

ALIGN_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
ALIGN_LEFT_WRAP = Alignment(horizontal='left', vertical='top', wrap_text=True)
ALIGN_LEFT = Alignment(horizontal='left', vertical='center')

THIN_BORDER = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9'),
)


# ═══════════════════════════════════════════════════
#  SHEET 1: Overview (matching OctaSales Overview)
# ═══════════════════════════════════════════════════
ws = wb.active
ws.title = "Overview"
ws.sheet_properties.tabColor = "0070C0"
ws.column_dimensions['A'].width = 3
ws.column_dimensions['B'].width = 25
ws.column_dimensions['C'].width = 65

ws.merge_cells('B2:C2')
ws['B2'] = "Refund Statement 1A Template"
ws['B2'].font = FONT_TITLE
ws['B2'].fill = FILL_TITLE
ws['B2'].alignment = Alignment(horizontal='left', vertical='center')

ws['B5'] = "Refund Type"
ws['C5'] = "ITC accumulated due to Inverted Tax Structure"
ws['B5'].font = FONT_WHITE_BOLD_11
ws['B5'].fill = FILL_REQUIRED
ws['C5'].font = FONT_NORMAL_11

ws['B6'] = "Legal Basis"
ws['C6'] = "Section 54(3) CGST Act, Rule 89(2)(h), Rule 89(5)"
ws['B6'].font = FONT_WHITE_BOLD_11
ws['B6'].fill = FILL_REQUIRED
ws['C6'].font = FONT_NORMAL_11

ws['B7'] = "Refund Code / Version"
ws['C7'] = "INVITC / 2.0"
ws['B7'].font = FONT_WHITE_BOLD_11
ws['B7'].fill = FILL_REQUIRED
ws['C7'].font = FONT_NORMAL_11

ws['B9'] = "Sheet Name"
ws['C9'] = "Description"
ws['B9'].font = FONT_WHITE_BOLD_11
ws['B9'].fill = FILL_REQUIRED
ws['C9'].font = FONT_WHITE_BOLD_11
ws['C9'].fill = FILL_REQUIRED

sheets_info = [
    ("Header", "GSTIN and Return Period details"),
    ("Inward", "Details of inward supplies received and ITC claimed"),
    ("Outward", "Details of outward supplies issued at inverted rate"),
    ("Help", "Field specifications, validation rules, and usage guidance"),
]
for i, (sname, desc) in enumerate(sheets_info):
    r = 10 + i
    ws[f'B{r}'] = sname
    ws[f'C{r}'] = desc
    ws[f'B{r}'].font = FONT_WHITE_BOLD_11
    ws[f'B{r}'].fill = FILL_REQUIRED
    ws[f'C{r}'].font = FONT_NORMAL_11

ws['B15'] = "Colour Coding"
ws['B15'].font = FONT_BOLD_11

ws.merge_cells('B16:C16')
ws['B16'] = "Mandatory Columns"
ws['B16'].font = FONT_WHITE_BOLD_11
ws['B16'].fill = FILL_REQUIRED

ws.merge_cells('B17:C17')
ws['B17'] = "Optional Columns"
ws['B17'].font = FONT_WHITE_BOLD_11
ws['B17'].fill = FILL_OPTIONAL

ws.merge_cells('B18:C18')
ws['B18'] = "Conditional Columns"
ws['B18'].font = FONT_WHITE_BOLD_11
ws['B18'].fill = FILL_CONDITIONAL

ws.merge_cells('B20:C20')
ws['B20'] = "Refund template for ITC accumulated due to Inverted Tax Structure"
ws['B20'].font = FONT_BOLD_11


# ═══════════════════════════════════════════════════
#  SHEET 2: Header
# ═══════════════════════════════════════════════════
ws_hd = wb.create_sheet("Header")
ws_hd.sheet_properties.tabColor = "0070C0"
ws_hd.column_dimensions['A'].width = 28
ws_hd.column_dimensions['B'].width = 30
ws_hd.column_dimensions['C'].width = 50

# Header row with OctaSales blue
for col, val in enumerate(["Field Name", "Value", "Description"], 1):
    cell = ws_hd.cell(row=1, column=col, value=val)
    cell.font = FONT_WHITE_BOLD
    cell.fill = FILL_REQUIRED
    cell.alignment = ALIGN_CENTER
    cell.border = THIN_BORDER

fields = [
    ("GSTIN", "", "15-character GSTIN of the refund applicant"),
    ("From Period (MM-YYYY)", "", "Starting return period e.g., 07-2024 for July 2024"),
    ("To Period (MM-YYYY)", "", "Ending return period e.g., 09-2024 for September 2024"),
]
for i, (field, val, desc) in enumerate(fields):
    r = i + 2
    ws_hd.cell(row=r, column=1, value=field).font = FONT_BOLD_10
    ws_hd.cell(row=r, column=2, value=val).font = FONT_NORMAL_10
    ws_hd.cell(row=r, column=3, value=desc).font = FONT_NORMAL_10
    for c in range(1, 4):
        ws_hd.cell(row=r, column=c).border = THIN_BORDER


# ═══════════════════════════════════════════════════
#  SHEET 3: Inward — colours on header row, data from row 2
# ═══════════════════════════════════════════════════
ws_in = wb.create_sheet("Inward")
ws_in.sheet_properties.tabColor = "00B050"

# (name, width, fill_category)
inward_cols = [
    ("Inward Supply Type", 30, FILL_REQUIRED),
    ("Supplier GSTIN", 18, FILL_CONDITIONAL),
    ("Doc Type", 14, FILL_REQUIRED),
    ("Doc No", 18, FILL_REQUIRED),
    ("Doc Date", 14, FILL_REQUIRED),
    ("Port Code", 10, FILL_CONDITIONAL),
    ("Taxable Value", 16, FILL_REQUIRED),
    ("IGST", 14, FILL_CONDITIONAL),
    ("CGST", 14, FILL_CONDITIONAL),
    ("SGST", 14, FILL_CONDITIONAL),
]

for ci, (name, width, fill) in enumerate(inward_cols):
    col = ci + 1
    ws_in.column_dimensions[get_column_letter(col)].width = width
    cell = ws_in.cell(row=1, column=col, value=name)
    cell.font = FONT_WHITE_BOLD
    cell.fill = fill
    cell.alignment = ALIGN_CENTER
    cell.border = THIN_BORDER

# TEXT format on text-type columns — preserves leading zeros, prevents
# Excel from converting invoice numbers like "001234" to the number 1234.
for text_col in ['B', 'D']:  # Supplier GSTIN, Doc No
    ws_in.column_dimensions[text_col].number_format = '@'

# Sample data at row 2 (no gap row)
sample = [
    "Inward Supply from Registered Person", "24AABCT1332L1ZX",
    "Invoice/Bill of Entry", "INV/2024/001", "15-08-2024", "", 500000, 90000, "", ""
]
for ci, val in enumerate(sample):
    cell = ws_in.cell(row=2, column=ci+1, value=val)
    cell.font = FONT_NORMAL_10
    cell.border = THIN_BORDER

# Dropdowns
dv_type = DataValidation(type="list",
    formula1='"Import of Goods/Supplies from SEZ to DTA,Import of Services/Supplies from SEZ to DTA,Inward supplies liable to reverse charge,Inward Supply from Registered Person,Inward Supplies from ISD"',
    allow_blank=True)
dv_type.error = "Select a valid Inward Supply Type"
ws_in.add_data_validation(dv_type)
dv_type.add('A2:A10001')

dv_doc = DataValidation(type="list", formula1='"Invoice/Bill of Entry,Debit Note,Credit Note"', allow_blank=True)
ws_in.add_data_validation(dv_doc)
dv_doc.add('C2:C10001')

ws_in.freeze_panes = 'A2'


# ═══════════════════════════════════════════════════
#  SHEET 4: Outward — same approach
# ═══════════════════════════════════════════════════
ws_out = wb.create_sheet("Outward")
ws_out.sheet_properties.tabColor = "E67E22"

outward_cols = [
    ("Outward Supply Type", 22, FILL_REQUIRED),
    ("Doc Type", 14, FILL_REQUIRED),
    ("Doc No", 18, FILL_CONDITIONAL),
    ("Doc Date", 14, FILL_CONDITIONAL),
    ("Taxable Value", 16, FILL_REQUIRED),
    ("IGST", 14, FILL_CONDITIONAL),
    ("CGST", 14, FILL_CONDITIONAL),
    ("SGST", 14, FILL_CONDITIONAL),
]

for ci, (name, width, fill) in enumerate(outward_cols):
    col = ci + 1
    ws_out.column_dimensions[get_column_letter(col)].width = width
    cell = ws_out.cell(row=1, column=col, value=name)
    cell.font = FONT_WHITE_BOLD
    cell.fill = fill
    cell.alignment = ALIGN_CENTER
    cell.border = THIN_BORDER

# TEXT format on text-type columns
for text_col in ['C']:  # Doc No
    ws_out.column_dimensions[text_col].number_format = '@'

sample_out = ["B2B", "Invoice/Bill of Entry", "SALE/2024/101", "20-08-2024", 200000, 10000, "", ""]
for ci, val in enumerate(sample_out):
    cell = ws_out.cell(row=2, column=ci+1, value=val)
    cell.font = FONT_NORMAL_10
    cell.border = THIN_BORDER

dv_otype = DataValidation(type="list", formula1='"B2B,B2C-Large,B2C-Small"', allow_blank=True)
ws_out.add_data_validation(dv_otype)
dv_otype.add('A2:A10001')

dv_odoc = DataValidation(type="list", formula1='"Invoice/Bill of Entry,Debit Note,Credit Note"', allow_blank=True)
ws_out.add_data_validation(dv_odoc)
dv_odoc.add('B2:B10001')

ws_out.freeze_panes = 'A2'


# ═══════════════════════════════════════════════════
#  SHEET 5: Help (exact OctaSales style)
# ═══════════════════════════════════════════════════
ws_h = wb.create_sheet("Help")
ws_h.sheet_properties.tabColor = "FFC000"
ws_h.column_dimensions['A'].width = 14
ws_h.column_dimensions['B'].width = 24
ws_h.column_dimensions['C'].width = 18
ws_h.column_dimensions['D'].width = 75
ws_h.column_dimensions['E'].width = 25

r = 2

# ─── Section 1: Header ───
ws_h.merge_cells(f'A{r}:E{r}')
ws_h[f'A{r}'] = "1. Sheet Name: Header"
ws_h[f'A{r}'].font = FONT_SECTION
ws_h[f'A{r}'].fill = FILL_SECTION_HEADER
r += 1

ws_h.merge_cells(f'A{r}:E{r}')
ws_h[f'A{r}'] = "The following table specifies the header fields for the refund application."
ws_h[f'A{r}'].font = FONT_NORMAL_11
r += 1

# Table header
for ci, h in enumerate(["Field Category", "Field Name", "Field Type", "Description", "Validation"]):
    cell = ws_h.cell(row=r, column=ci+1, value=h)
    cell.font = FONT_HELP_TABLE_HEADER
    cell.border = THIN_BORDER
r += 1

def help_row(ws, row, cat, name, ftype, desc, validation=""):
    fills = {"Required": FILL_REQUIRED, "Optional": FILL_OPTIONAL, "Conditional": FILL_CONDITIONAL}
    vals = [cat, name, ftype, desc, validation]
    for ci, val in enumerate(vals):
        cell = ws.cell(row=row, column=ci+1, value=val)
        cell.font = FONT_NORMAL_10
        cell.border = THIN_BORDER
        cell.alignment = ALIGN_LEFT_WRAP
        if ci == 0 and cat in fills:
            cell.fill = fills[cat]
            cell.font = Font(name='Calibri', size=10, color='FFFFFF')
    return row + 1

r = help_row(ws_h, r, "Required", "GSTIN", "Text (15)", "15-character GSTIN of the refund applicant.", "GSTIN regex; auto upper-cased and trimmed")
r = help_row(ws_h, r, "Required", "From Period", "Text (MM-YYYY)", "Starting return period.\nExample: 07-2024 for July 2024", "Must be >= 07-2017; <= current; <= To Period")
r = help_row(ws_h, r, "Required", "To Period", "Text (MM-YYYY)", "Ending return period.\nExample: 09-2024 for September 2024", "Must be >= 07-2017; <= current; >= From Period")
r += 1

# ─── Section 2: Inward ───
ws_h.merge_cells(f'A{r}:E{r}')
ws_h[f'A{r}'] = "2. Sheet Name: Inward"
ws_h[f'A{r}'].font = FONT_SECTION
ws_h[f'A{r}'].fill = FILL_SECTION_HEADER
r += 1

ws_h.merge_cells(f'A{r}:E{r}')
ws_h[f'A{r}'] = "The following table specifies the fields for inward supply documents and ITC claimed."
ws_h[f'A{r}'].font = FONT_NORMAL_11
r += 1

for ci, h in enumerate(["Field Category", "Field Name", "Field Type", "Description", "Validation"]):
    cell = ws_h.cell(row=r, column=ci+1, value=h)
    cell.font = FONT_HELP_TABLE_HEADER
    cell.border = THIN_BORDER
r += 1

r = help_row(ws_h, r, "Required", "Inward Supply Type", "Text",
    "Type of inward supply.\nValid values:\n  Import of Goods/Supplies from SEZ to DTA\n  Import of Services/Supplies from SEZ to DTA\n  Inward supplies liable to reverse charge\n  Inward Supply from Registered Person\n  Inward Supplies from ISD",
    "Dropdown provided")
r = help_row(ws_h, r, "Conditional", "Supplier GSTIN", "Text (15)",
    "GSTIN of supplier or ISD.\nRequired for: Registered Person, ISD.\nAuto-filled by app for: Import, RCM (= self GSTIN from Header).\nMust differ from applicant GSTIN.",
    "GSTIN regex; must differ from self")
r = help_row(ws_h, r, "Required", "Doc Type", "Text",
    "Type of document.\nValid values: Invoice/Bill of Entry, Debit Note, Credit Note\nFor Import of Goods, only Invoice/Bill of Entry is allowed.",
    "Dropdown provided")
r = help_row(ws_h, r, "Required", "Doc No", "Text (16)",
    "Document number or Bill of Entry number.\nMax 16 alphanumeric characters, / and - allowed.",
    "Cannot be 0 or special chars only")
r = help_row(ws_h, r, "Required", "Doc Date", "Text (DD-MM-YYYY)",
    "Date of document.\nMust be on or after 01-07-2017.\nCannot be a future date.",
    "Valid calendar date; <= To Period")
r = help_row(ws_h, r, "Conditional", "Port Code", "Text (6)",
    "Port code.\nRequired ONLY for: Import of Goods/Supplies from SEZ to DTA.\nLeave blank for all other inward types.",
    "Exactly 6 alphanumeric characters")
r = help_row(ws_h, r, "Required", "Taxable Value", "Number (15,2)",
    "Taxable value of the inward supply.\nMax 15 digits before decimal, 2 after.",
    "Must be >= 0")
r = help_row(ws_h, r, "Conditional", "IGST", "Number (15,2)",
    "Integrated Tax amount.\nImport types: IGST only (CGST/SGST must be blank).\nRCM/Registered: IGST alone OR CGST+SGST (not both).\nISD: all three allowed simultaneously.",
    "Non-negative; mutual exclusivity per type")
r = help_row(ws_h, r, "Conditional", "CGST", "Number (15,2)",
    "Central Tax amount.\nNot applicable for Import types.\nFor RCM/Registered: required if no IGST.\nFor ISD: allowed alongside IGST.",
    "Non-negative; paired with SGST")
r = help_row(ws_h, r, "Conditional", "SGST", "Number (15,2)",
    "State/UT Tax amount.\nSame rules as CGST.\nIf CGST is provided, SGST must also be provided and vice versa.",
    "Non-negative; paired with CGST")
r += 1

# ─── Section 3: Outward ───
ws_h.merge_cells(f'A{r}:E{r}')
ws_h[f'A{r}'] = "3. Sheet Name: Outward"
ws_h[f'A{r}'].font = FONT_SECTION
ws_h[f'A{r}'].fill = FILL_SECTION_HEADER
r += 1

ws_h.merge_cells(f'A{r}:E{r}')
ws_h[f'A{r}'] = "The following table specifies the fields for outward supply documents at inverted rate."
ws_h[f'A{r}'].font = FONT_NORMAL_11
r += 1

for ci, h in enumerate(["Field Category", "Field Name", "Field Type", "Description", "Validation"]):
    cell = ws_h.cell(row=r, column=ci+1, value=h)
    cell.font = FONT_HELP_TABLE_HEADER
    cell.border = THIN_BORDER
r += 1

r = help_row(ws_h, r, "Required", "Outward Supply Type", "Text",
    "Type of outward supply.\nValid values: B2B, B2C-Large, B2C-Small",
    "Dropdown provided")
r = help_row(ws_h, r, "Required", "Doc Type", "Text",
    "Type of document.\nValid values: Invoice/Bill of Entry, Debit Note, Credit Note",
    "Dropdown provided")
r = help_row(ws_h, r, "Conditional", "Doc No", "Text (16)",
    "Document number.\nFor B2C-Small: leave blank. App auto-fills 'B2COTH' in JSON.\nFor B2B and B2C-Large: mandatory.",
    "Max 16 alphanumeric + / and -")
r = help_row(ws_h, r, "Conditional", "Doc Date", "Text (DD-MM-YYYY)",
    "Date of document.\nFor B2C-Small: leave blank. App auto-fills '01-07-2017' in JSON.\nFor B2B and B2C-Large: mandatory.",
    "Valid calendar date; cannot be future")
r = help_row(ws_h, r, "Required", "Taxable Value", "Number (15,2)",
    "Taxable value of the outward supply.\nMax 15 digits before decimal, 2 after.",
    "Must be >= 0")
r = help_row(ws_h, r, "Conditional", "IGST", "Number (15,2)",
    "Integrated Tax amount.\nB2B: IGST alone OR CGST+SGST (not both).\nB2C-Large: IGST only (mandatory).\nB2C-Small: not applicable (leave blank).",
    "Non-negative; <= Taxable Value")
r = help_row(ws_h, r, "Conditional", "CGST", "Number (15,2)",
    "Central Tax amount.\nB2B: required if no IGST.\nB2C-Large: not applicable.\nB2C-Small: mandatory (with SGST).",
    "Non-negative; paired with SGST")
r = help_row(ws_h, r, "Conditional", "SGST", "Number (15,2)",
    "State/UT Tax amount.\nSame rules as CGST. Always paired with CGST.",
    "Non-negative; paired with CGST")
r += 1

# ─── Section 4: Tax Rules ───
ws_h.merge_cells(f'A{r}:E{r}')
ws_h[f'A{r}'] = "4. Tax Mutual Exclusivity Rules"
ws_h[f'A{r}'].font = FONT_SECTION
ws_h[f'A{r}'].fill = FILL_SECTION_HEADER
r += 1
r += 1

rules = [
    "INWARD — Import of Goods / Import of Services:  IGST only. CGST and SGST must be blank.",
    "INWARD — RCM / Registered Person:  Either IGST alone OR (CGST + SGST) together. Cannot fill all three.",
    "INWARD — ISD:  IGST, CGST, and SGST are all allowed simultaneously.",
    "OUTWARD — B2B:  Either IGST alone OR (CGST + SGST) together. Cannot fill all three.",
    "OUTWARD — B2C-Large:  IGST only. CGST and SGST must be blank.",
    "OUTWARD — B2C-Small:  CGST + SGST only. IGST must be blank. Doc No and Date auto-filled by app.",
    "ALL:  Sum of tax (IGST or CGST+SGST) must be <= Taxable Value.",
]
for rule in rules:
    ws_h.merge_cells(f'A{r}:E{r}')
    ws_h[f'A{r}'] = f"• {rule}"
    ws_h[f'A{r}'].font = FONT_NORMAL_10
    r += 1
r += 1

# ─── Section 5: Duplicate Detection ───
ws_h.merge_cells(f'A{r}:E{r}')
ws_h[f'A{r}'] = "5. Duplicate Detection"
ws_h[f'A{r}'].font = FONT_SECTION
ws_h[f'A{r}'].fill = FILL_SECTION_HEADER
r += 1
r += 1

dups = [
    "INWARD:  Duplicate key = Supplier GSTIN + Doc No + Financial Year + Month.",
    "OUTWARD:  Duplicate key = Outward Type + Doc Type + Doc No + Financial Year + Month. DN/CN grouped as 'NOTE'.",
    "Financial year adjustment: If invoice month is Jan-Mar (1-3), FY year is decremented by 1.",
    "B2C-Small rows are excluded from duplicate checking (they share the auto-filled 'B2COTH' doc no).",
]
for d in dups:
    ws_h.merge_cells(f'A{r}:E{r}')
    ws_h[f'A{r}'] = f"• {d}"
    ws_h[f'A{r}'].font = FONT_NORMAL_10
    r += 1
r += 1

# ─── Section 6: Running the Tool ───
ws_h.merge_cells(f'A{r}:E{r}')
ws_h[f'A{r}'] = "6. Running the Tool"
ws_h[f'A{r}'].font = FONT_SECTION
ws_h[f'A{r}'].fill = FILL_SECTION_HEADER
r += 1
r += 1

cmds = [
    "python main.py                         → Opens file picker",
    "python main.py input.xlsx              → Process specific file",
    "python main.py input.xlsx --pretty     → Pretty-print JSON output",
    "python main.py input.xlsx -v           → Verbose validation output",
]
for cmd in cmds:
    ws_h.merge_cells(f'A{r}:E{r}')
    ws_h[f'A{r}'] = f"  {cmd}"
    ws_h[f'A{r}'].font = Font(name='Consolas', size=10)
    r += 1

ws_h.freeze_panes = 'A3'

# ─── Set active to Overview ───
wb.active = wb.sheetnames.index("Overview")

apply_octa_replica_styling(wb)
wb.save(os.path.join(os.getcwd(), "GST_Refund_Stmt1A_Template.xlsx"))
output = os.path.join(os.getcwd(), "GST_Refund_Stmt1A_Template.xlsx")
print(f"Created: {output}")
