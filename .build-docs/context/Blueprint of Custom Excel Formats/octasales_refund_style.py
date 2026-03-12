from __future__ import annotations

from copy import copy
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

BLUE = 'FF0070C0'
GREEN = 'FF00B050'
CYAN = 'FF00B0F0'
AMBER = 'FFFFC000'
ORANGE = 'FFED7D31'
LIGHT_GREY = 'FFD9D9D9'
BLACK = 'FF000000'
WHITE = 'FFFFFFFF'

THIN_BLACK = Border(
    left=Side(style='thin', color='FF000000'),
    right=Side(style='thin', color='FF000000'),
    top=Side(style='thin', color='FF000000'),
    bottom=Side(style='thin', color='FF000000'),
)

FONT_TITLE = Font(name='Calibri', size=20, bold=True, color=WHITE)
FONT_WHITE_11 = Font(name='Calibri', size=11, bold=True, color=WHITE)
FONT_WHITE_10 = Font(name='Calibri', size=10, bold=True, color=WHITE)
FONT_BLACK_14 = Font(name='Calibri', size=14, bold=True)
FONT_BLACK_11 = Font(name='Calibri', size=11, bold=True)
FONT_BLACK_10 = Font(name='Calibri', size=10)
FONT_FOOTER = Font(name='Calibri', size=11, bold=True, color=WHITE)

FILL_BLUE = PatternFill('solid', fgColor=BLUE)
FILL_GREEN = PatternFill('solid', fgColor=GREEN)
FILL_CYAN = PatternFill('solid', fgColor=CYAN)
FILL_AMBER = PatternFill('solid', fgColor=AMBER)
FILL_ORANGE = PatternFill('solid', fgColor=ORANGE)
FILL_LIGHT_GREY = PatternFill('solid', fgColor=LIGHT_GREY)

ALIGN_LEFT_CENTER = Alignment(horizontal='left', vertical='center')
ALIGN_LEFT_WRAP = Alignment(horizontal='left', vertical='center', wrap_text=True)
ALIGN_TITLE = Alignment(horizontal='left', vertical='center')

DONOR_CANDIDATES = (
    'OctaSales (2).xlsx',
    'OctaSales.xlsx',
)


def _load_donor() -> Optional[openpyxl.Workbook]:
    base = Path(__file__).resolve().parent
    for name in DONOR_CANDIDATES:
        path = base / name
        if path.exists():
            try:
                return openpyxl.load_workbook(path)
            except Exception:
                pass
    return None


_DONOR = _load_donor()


def _copy_style(src, dst):
    dst.font = copy(src.font)
    dst.fill = copy(src.fill)
    dst.border = copy(src.border)
    dst.alignment = copy(src.alignment)
    dst.number_format = src.number_format
    dst.protection = copy(src.protection)
    return dst


def _fill_rgb(cell) -> str:
    try:
        return (cell.fill.fgColor.rgb or '').upper()
    except Exception:
        return ''


def _style_data_header(dst, kind: str):
    if _DONOR is not None:
        donor = _DONOR['Sales']['A1'] if kind == 'required' else _DONOR['Sales']['B1'] if kind == 'optional' else _DONOR['Sales']['L1']
        _copy_style(donor, dst)
        return
    dst.font = copy(FONT_WHITE_10)
    dst.fill = copy(FILL_BLUE if kind == 'required' else FILL_GREEN if kind == 'optional' else FILL_CYAN)
    dst.alignment = copy(ALIGN_LEFT_CENTER)
    dst.border = Border()


def _style_plain_data(dst):
    if _DONOR is not None:
        donor = _DONOR['Sales']['A2']
        _copy_style(donor, dst)
        return
    dst.font = Font(name='Calibri', size=11)
    dst.alignment = Alignment(horizontal='general', vertical='bottom')
    dst.fill = PatternFill(fill_type=None)
    dst.border = Border()


def _style_overview_title(dst):
    if _DONOR is not None:
        _copy_style(_DONOR['Overview']['B2'], dst)
        return
    dst.font = copy(FONT_TITLE)
    dst.fill = copy(FILL_BLUE)
    dst.alignment = copy(ALIGN_TITLE)
    dst.border = copy(THIN_BLACK)


def _style_overview_block(dst):
    if _DONOR is not None:
        _copy_style(_DONOR['Overview']['B4'], dst)
        return
    dst.fill = copy(FILL_BLUE)
    dst.border = copy(THIN_BLACK)
    dst.font = copy(FONT_TITLE)
    dst.alignment = copy(ALIGN_TITLE)


def _style_overview_label(dst):
    if _DONOR is not None:
        _copy_style(_DONOR['Overview']['B5'], dst)
        return
    dst.font = copy(FONT_WHITE_11)
    dst.fill = copy(FILL_BLUE)
    dst.alignment = copy(ALIGN_LEFT_CENTER)
    dst.border = copy(THIN_BLACK)


def _style_overview_value(dst):
    if _DONOR is not None:
        _copy_style(_DONOR['Overview']['C5'], dst)
        return
    dst.font = Font(name='Calibri', size=11)
    dst.alignment = copy(ALIGN_LEFT_CENTER)
    dst.fill = PatternFill(fill_type=None)
    dst.border = Border()


def _style_overview_table_head_b(dst):
    if _DONOR is not None:
        _copy_style(_DONOR['Overview']['B9'], dst)
        return
    dst.font = copy(FONT_WHITE_11)
    dst.fill = copy(FILL_BLUE)
    dst.alignment = copy(ALIGN_LEFT_CENTER)
    dst.border = copy(THIN_BLACK)


def _style_overview_table_head_c(dst):
    if _DONOR is not None:
        _copy_style(_DONOR['Overview']['C9'], dst)
        return
    dst.font = copy(FONT_WHITE_11)
    dst.fill = copy(FILL_BLUE)
    dst.alignment = copy(ALIGN_LEFT_CENTER)
    dst.border = copy(THIN_BLACK)


def _style_overview_sheet_name(dst):
    if _DONOR is not None:
        _copy_style(_DONOR['Overview']['B10'], dst)
        return
    dst.font = copy(FONT_WHITE_11)
    dst.fill = copy(FILL_BLUE)
    dst.alignment = copy(ALIGN_LEFT_CENTER)
    dst.border = copy(THIN_BLACK)


def _style_overview_desc(dst, wrap=False):
    if _DONOR is not None:
        donor = _DONOR['Overview']['C11'] if wrap else _DONOR['Overview']['C10']
        _copy_style(donor, dst)
        return
    dst.font = Font(name='Calibri', size=11)
    dst.alignment = copy(ALIGN_LEFT_WRAP if wrap else ALIGN_LEFT_CENTER)
    dst.border = copy(THIN_BLACK)


def _style_overview_colour_heading(dst):
    if _DONOR is not None:
        _copy_style(_DONOR['Overview']['B17'], dst)
        return
    dst.font = copy(FONT_BLACK_11)
    dst.alignment = copy(ALIGN_LEFT_CENTER)
    dst.border = Border()


def _style_overview_legend(dst, kind: str):
    if _DONOR is not None:
        donor = _DONOR['Overview']['B18'] if kind == 'required' else _DONOR['Overview']['B19'] if kind == 'optional' else _DONOR['Overview']['B20']
        _copy_style(donor, dst)
        return
    dst.font = copy(FONT_WHITE_11)
    dst.fill = copy(FILL_BLUE if kind == 'required' else FILL_GREEN if kind == 'optional' else FILL_CYAN)
    dst.alignment = copy(ALIGN_LEFT_CENTER)
    dst.border = Border()


def _style_overview_footer(dst):
    if _DONOR is not None:
        _copy_style(_DONOR['Overview']['B22'], dst)
        return
    dst.font = copy(FONT_FOOTER)
    dst.fill = copy(FILL_ORANGE)
    dst.alignment = copy(ALIGN_LEFT_CENTER)
    dst.border = Border()


def _style_help_section(dst):
    if _DONOR is not None:
        _copy_style(_DONOR['Help']['A2'], dst)
        return
    dst.font = copy(FONT_BLACK_14)
    dst.fill = copy(FILL_AMBER)
    dst.alignment = copy(ALIGN_LEFT_CENTER)
    dst.border = copy(THIN_BLACK)


def _style_help_intro(dst, wrap=False):
    if _DONOR is not None:
        donor = _DONOR['Help']['A58'] if wrap else _DONOR['Help']['A3']
        _copy_style(donor, dst)
        return
    dst.font = Font(name='Calibri', size=11)
    dst.alignment = copy(ALIGN_LEFT_WRAP if wrap else ALIGN_LEFT_CENTER)
    dst.border = copy(THIN_BLACK)
    dst.fill = PatternFill(fill_type=None)


def _style_help_header(dst):
    if _DONOR is not None:
        _copy_style(_DONOR['Help']['A4'], dst)
        return
    dst.font = copy(FONT_BLACK_11)
    dst.fill = copy(FILL_LIGHT_GREY)
    dst.alignment = copy(ALIGN_LEFT_CENTER)
    dst.border = copy(THIN_BLACK)


_HELP_COND_DONOR = None
if _DONOR is not None:
    hws = _DONOR['Help']
    for row in range(5, hws.max_row + 1):
        if str(hws.cell(row, 1).value or '').strip().lower() == 'conditional':
            _HELP_COND_DONOR = hws.cell(row, 1)
            break


def _style_help_category(dst, kind: str):
    if _DONOR is not None:
        donor = _DONOR['Help']['A5'] if kind == 'required' else _DONOR['Help']['A6'] if kind == 'optional' else (_HELP_COND_DONOR or _DONOR['Help']['A5'])
        _copy_style(donor, dst)
        return
    dst.font = Font(name='Calibri', size=10, color=WHITE)
    dst.fill = copy(FILL_BLUE if kind == 'required' else FILL_GREEN if kind == 'optional' else FILL_CYAN)
    dst.alignment = copy(ALIGN_LEFT_CENTER)
    dst.border = copy(THIN_BLACK)


def _style_help_body(dst):
    if _DONOR is not None:
        _copy_style(_DONOR['Help']['B5'], dst)
        return
    dst.font = copy(FONT_BLACK_10)
    dst.alignment = copy(ALIGN_LEFT_WRAP)
    dst.border = copy(THIN_BLACK)


def _setup_sheet_defaults(ws):
    ws.sheet_properties.tabColor = None
    ws.page_setup.paperSize = 9
    ws.page_setup.orientation = 'portrait'
    ws.page_margins.left = 0.7
    ws.page_margins.right = 0.7
    ws.page_margins.top = 0.75
    ws.page_margins.bottom = 0.75


_OVERVIEW_WIDTHS = {'A': 8.0, 'B': 19.1640625, 'C': 90.83203125}
_HELP_WIDTHS = {'A': 14.0, 'B': 21.1640625, 'C': 12.6640625, 'D': 53.1640625, 'E': 18.5, 'F': 58.5}
_OVERVIEW_ROW_HEIGHTS = {
    1: 8.75, 2: 28.5, 3: 8.0, 4: 10.25, 5: 15.0, 6: 15.0, 7: 15.0,
    9: 28.5, 10: 15.0, 11: 32.0, 12: 17.25, 13: 17.25, 14: 17.25,
    15: 20.5, 16: 15.0, 17: 15.0, 18: 15.0, 19: 15.0, 20: 15.0,
    21: 15.0, 22: 15.0,
}


def _ensure_merge(ws, rng: str):
    if rng not in [str(r) for r in ws.merged_cells.ranges]:
        ws.merge_cells(rng)


def _restyle_overview(ws):
    _setup_sheet_defaults(ws)
    ws.sheet_format.defaultRowHeight = 15.0
    for col, width in _OVERVIEW_WIDTHS.items():
        ws.column_dimensions[col].width = width
    for row, height in _OVERVIEW_ROW_HEIGHTS.items():
        ws.row_dimensions[row].height = height

    _ensure_merge(ws, 'B2:C2')
    _style_overview_title(ws['B2'])

    if ws['B4'].value is None and ws['C4'].value is None:
        try:
            _ensure_merge(ws, 'B4:C4')
        except Exception:
            pass
        _style_overview_block(ws['B4'])

    nonempty_b = {r: str(ws[f'B{r}'].value).strip() for r in range(1, ws.max_row + 1) if ws[f'B{r}'].value not in (None, '')}
    sheet_name_row = next((r for r, v in nonempty_b.items() if v == 'Sheet Name'), None)
    colour_row = next((r for r, v in nonempty_b.items() if v == 'Colour Coding'), None)

    if sheet_name_row:
        for r in range(5, sheet_name_row):
            if ws[f'B{r}'].value not in (None, ''):
                _style_overview_label(ws[f'B{r}'])
            if ws[f'C{r}'].value not in (None, ''):
                _style_overview_value(ws[f'C{r}'])
        _style_overview_table_head_b(ws[f'B{sheet_name_row}'])
        _style_overview_table_head_c(ws[f'C{sheet_name_row}'])
        end_row = (colour_row - 2) if colour_row else ws.max_row
        for r in range(sheet_name_row + 1, end_row + 1):
            if ws[f'B{r}'].value not in (None, ''):
                _style_overview_sheet_name(ws[f'B{r}'])
            if ws[f'C{r}'].value not in (None, ''):
                text = str(ws[f'C{r}'].value)
                _style_overview_desc(ws[f'C{r}'], wrap=(len(text) > 95 or '\n' in text))
                if len(text) > 110 and ws.row_dimensions[r].height in (None, 15.0):
                    ws.row_dimensions[r].height = 32.0

    if colour_row:
        _style_overview_colour_heading(ws[f'B{colour_row}'])
        legend_rows = []
        for r in range(colour_row + 1, min(colour_row + 6, ws.max_row + 1)):
            txt = str(ws[f'B{r}'].value or '').lower()
            if 'mandatory' in txt:
                legend_rows.append((r, 'required'))
            elif 'optional' in txt:
                legend_rows.append((r, 'optional'))
            elif 'conditional' in txt:
                legend_rows.append((r, 'conditional'))
        for r, kind in legend_rows:
            try:
                _ensure_merge(ws, f'B{r}:C{r}')
            except Exception:
                pass
            _style_overview_legend(ws[f'B{r}'], kind)

        footer_row = next((r for r in range(colour_row + 1, ws.max_row + 1) if ws[f'B{r}'].value not in (None, '') and 'columns' not in str(ws[f'B{r}'].value).lower()), None)
        if footer_row:
            for r in range(footer_row, ws.max_row + 1):
                if ws[f'B{r}'].value not in (None, '') and 'columns' not in str(ws[f'B{r}'].value).lower():
                    footer_row = r
            try:
                _ensure_merge(ws, f'B{footer_row}:C{footer_row}')
            except Exception:
                pass
            _style_overview_footer(ws[f'B{footer_row}'])


def _guess_kind_from_fill(cell) -> str:
    rgb = _fill_rgb(cell)
    if rgb == GREEN:
        return 'optional'
    if rgb == CYAN:
        return 'conditional'
    if rgb == BLUE:
        return 'required'
    return 'required'


def _restyle_data_like_sheet(ws):
    _setup_sheet_defaults(ws)
    ws.sheet_format.defaultRowHeight = 14.0
    ws.row_dimensions[1].height = 14.0
    if ws.max_row >= 2 and ws.freeze_panes is None:
        ws.freeze_panes = 'A2'

    for cell in ws[1]:
        if cell.value not in (None, ''):
            _style_data_header(cell, _guess_kind_from_fill(cell))

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            if cell.value not in (None, ''):
                _style_plain_data(cell)


_TITLE_PREFIXES = tuple(f'{i}. Sheet Name:' for i in range(1, 10))


def _restyle_help(ws):
    _setup_sheet_defaults(ws)
    ws.sheet_format.defaultRowHeight = 15.0
    for col, width in _HELP_WIDTHS.items():
        ws.column_dimensions[col].width = width

    for r in range(1, ws.max_row + 1):
        a = ws[f'A{r}']
        b = ws[f'B{r}']
        txt = str(a.value or '').strip()
        merged = any(m.min_row == r and m.min_col == 1 for m in ws.merged_cells.ranges)

        if merged and txt.startswith(_TITLE_PREFIXES):
            _style_help_section(a)
            if ws.row_dimensions[r].height in (None, 15.0):
                ws.row_dimensions[r].height = 19.0
            continue

        if merged and txt:
            wrap = ('\n' in txt) or len(txt) > 90
            _style_help_intro(a, wrap=wrap)
            if ws.row_dimensions[r].height in (None, 15.0):
                ws.row_dimensions[r].height = 30.0 if len(txt) < 120 else 60.0
            continue

        if txt == 'Field Category':
            for col in range(1, 6):
                _style_help_header(ws.cell(r, col))
            if ws.row_dimensions[r].height in (None, 15.0):
                ws.row_dimensions[r].height = 26.0
            continue

        row_has_content = any(ws.cell(r, col).value not in (None, '') for col in range(1, 6))
        if not row_has_content:
            continue

        category = str(a.value or '').strip().lower()
        if category in {'required', 'optional', 'conditional'}:
            _style_help_category(a, category)
        elif a.value not in (None, ''):
            _style_help_body(a)
        for col in range(2, 6):
            c = ws.cell(r, col)
            if c.value not in (None, ''):
                _style_help_body(c)


def apply_octa_replica_styling(wb: openpyxl.Workbook) -> openpyxl.Workbook:
    for ws in wb.worksheets:
        name = ws.title.strip().lower()
        if name == 'overview':
            _restyle_overview(ws)
        elif name == 'help':
            _restyle_help(ws)
        else:
            _restyle_data_like_sheet(ws)
    return wb
