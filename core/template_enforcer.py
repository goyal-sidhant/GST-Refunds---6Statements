"""
FILE: core/template_enforcer.py

PURPOSE: Enforces strict column header matching between the Excel file and
         the expected StatementConfig. If any column is missing, renamed,
         or in the wrong position, processing is immediately halted with
         a clear error message. Also verifies required sheets exist.

CONTAINS:
- enforce_template()  — Main enforcement function

DEPENDS ON:
- models/statement_config.py → StatementConfig, SheetConfig
- models/validation_result.py → ValidationResult
- config/error_messages.py → TEMPLATE_ERRORS

USED BY:
- readers/template_reader.py → called before reading any data rows

CHANGE LOG:
| Date       | Change                              | Why                                |
|------------|-------------------------------------|------------------------------------|
| 11-03-2026 | Created — strict template enforcer  | Phase 0 infrastructure setup       |
"""

# Third-party libraries
import openpyxl
from openpyxl.workbook import Workbook

# This project's modules
from models.statement_config import StatementConfig
from models.validation_result import ValidationResult
from config.error_messages import TEMPLATE_ERRORS


def enforce_template(
    workbook: Workbook,
    config: StatementConfig,
    result: ValidationResult,
) -> bool:
    """
    WHAT:
        Checks that the workbook has the correct sheets and that each
        data sheet has the exact expected column headers in the correct
        order (row 1). Adds errors to the ValidationResult if any
        mismatch is found.

    WHY ADDED:
        Rebuild Brief Section 3: "The tool must rigidly enforce this
        mapping. If any column is missing, renamed, or reordered, the
        tool must immediately reject the file."

    CALLED BY:
        → readers/template_reader.py → enforce_template() before reading data

    CALLS:
        → config/error_messages.py → TEMPLATE_ERRORS for message templates

    EDGE CASES HANDLED:
        - Sheet name is case-sensitive (must match exactly)
        - Extra columns after the expected ones are allowed (ignored)
        - Completely empty header row is caught as "column missing"
        - Header sheet is checked for existence separately

    ASSUMPTIONS:
        *** ASSUMPTION: Column headers are always in row 1 of each data sheet.
            The custom templates are designed this way. If a template
            has headers in a different row, this function will fail.
            Decided in Phase 0 design. ***

    PARAMETERS:
        workbook (Workbook):       The opened openpyxl workbook.
        config (StatementConfig):  The statement configuration to enforce against.
        result (ValidationResult): Errors are added here.

    RETURNS:
        bool: True if the template passes enforcement, False if errors were added.
    """
    is_valid = True

    # --- Check Header sheet exists ---
    if "Header" not in workbook.sheetnames:
        result.add_error(
            message=TEMPLATE_ERRORS["header_sheet_missing"],
            sheet="Header",
            category="template",
        )
        return False

    # --- Check each data sheet exists and has correct columns ---
    for sheet_config in config.sheets:
        if sheet_config.name not in workbook.sheetnames:
            result.add_error(
                message=TEMPLATE_ERRORS["sheet_missing"].format(
                    sheet_name=sheet_config.name,
                    found_sheets=", ".join(workbook.sheetnames),
                    stmt_type=config.code,
                ),
                sheet=sheet_config.name,
                category="template",
            )
            is_valid = False
            continue

        worksheet = workbook[sheet_config.name]
        expected_columns = sheet_config.columns

        # Read row 1 as headers
        header_row = []
        for col_idx in range(1, len(expected_columns) + 1):
            cell_value = worksheet.cell(row=1, column=col_idx).value
            header_row.append(str(cell_value).strip() if cell_value is not None else "")

        # Check column count
        if len(header_row) < len(expected_columns):
            result.add_error(
                message=TEMPLATE_ERRORS["column_missing"].format(
                    sheet_name=sheet_config.name,
                    column_name=expected_columns[len(header_row)].header,
                    expected_count=len(expected_columns),
                    found_count=len(header_row),
                ),
                sheet=sheet_config.name,
                category="template",
            )
            is_valid = False
            continue

        # Check each column header matches exactly
        for col_idx, column_spec in enumerate(expected_columns):
            actual_header = header_row[col_idx]
            expected_header = column_spec.header

            if actual_header != expected_header:
                result.add_error(
                    message=TEMPLATE_ERRORS["column_mismatch"].format(
                        sheet_name=sheet_config.name,
                        position=col_idx + 1,
                        expected=expected_header,
                        found=actual_header if actual_header else "(empty)",
                    ),
                    sheet=sheet_config.name,
                    field=expected_header,
                    category="template",
                )
                is_valid = False

    return is_valid
