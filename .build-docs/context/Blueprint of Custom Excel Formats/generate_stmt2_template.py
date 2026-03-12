#!/usr/bin/env python3
"""
GST Refund Offline Utility — Template Generator
Statement 2 — Export of Services with Payment of Tax (EXPWP)

Generates: GST_Refund_Stmt2_Template.xlsx
Usage:     python generate_stmt2_template.py
Requires:  pip install openpyxl

The generated template follows OctaSales design language:
  - Blue headers (0070C0) = Mandatory
  - Green headers (00B050) = Optional  
  - Light blue headers (00B0F0) = Conditional
  - Calibri font throughout
  - Data starts at row 2 (no gap row)
  - Dropdowns on type/classification columns
"""
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

wb = openpyxl.Workbook()

# ─── OctaSales colour scheme ───
FILL_REQ = PatternFill('solid', fgColor='0070C0')
FILL_OPT = PatternFill('solid', fgColor='00B050')
FILL_COND = PatternFill('solid', fgColor='00B0F0')
FILL_SECTION = PatternFill('solid', fgColor='FFC000')
FILL_TITLE = PatternFill('solid', fgColor='0070C0')

FW = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
FW11 = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
FB11 = Font(name='Calibri', bold=True, size=11)
FB10 = Font(name='Calibri', bold=True, size=10)
FN11 = Font(name='Calibri', size=11)
FN10 = Font(name='Calibri', size=10)
FTITLE = Font(name='Calibri', bold=True, color='FFFFFF', size=20)
FSEC = Font(name='Calibri', bold=True, size=14)
FHELP_H = Font(name='Calibri', bold=True, size=11)

AC = Alignment(horizontal='center', vertical='center', wrap_text=True)
ALW = Alignment(horizontal='left', vertical='top', wrap_text=True)
TB = Border(left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9'))

def help_row(ws, row, cat, name, ftype, desc, val=""):
    fills = {"Required": FILL_REQ, "Optional": FILL_OPT, "Conditional": FILL_COND}
    for ci, v in enumerate([cat, name, ftype, desc, val]):
        cell = ws.cell(row=row, column=ci+1, value=v)
        cell.font = FN10
        cell.border = TB
        cell.alignment = ALW
        if ci == 0 and cat in fills:
            cell.fill = fills[cat]
            cell.font = Font(name='Calibri', size=10, color='FFFFFF')
    return row + 1

# ═══════════════════════════════════════
#  Overview
# ═══════════════════════════════════════
ws = wb.active
ws.title = "Overview"
ws.sheet_properties.tabColor = "0070C0"
ws.column_dimensions['A'].width = 3
ws.column_dimensions['B'].width = 25
ws.column_dimensions['C'].width = 65

ws.merge_cells('B2:C2')
ws['B2'] = "Refund Statement 2 Template"
ws['B2'].font = FTITLE
ws['B2'].fill = FILL_TITLE

ws['B5'] = "Refund Type"
ws['C5'] = "Export of Services — with Payment of Tax"
ws['B6'] = "Legal Basis"
ws['C6'] = "Section 54(3) CGST Act, Rule 89(2)(c), Rule 89(4)"
ws['B7'] = "Refund Code / Version"
ws['C7'] = "EXPWP / 3.0"
for r in range(5, 8):
    ws[f'B{r}'].font = FW11
    ws[f'B{r}'].fill = FILL_REQ
    ws[f'C{r}'].font = FN11

ws['B9'] = "Sheet Name"
ws['C9'] = "Description"
ws['B9'].font = FW11; ws['B9'].fill = FILL_REQ
ws['C9'].font = FW11; ws['C9'].fill = FILL_REQ

for i, (s, d) in enumerate([
    ("Header", "GSTIN of the refund applicant"),
    ("Data", "Export service invoices, tax paid, and BRC/FIRC details"),
    ("Help", "Field specifications, validation rules, and usage guidance"),
]):
    ws[f'B{10+i}'] = s; ws[f'C{10+i}'] = d
    ws[f'B{10+i}'].font = FW11; ws[f'B{10+i}'].fill = FILL_REQ
    ws[f'C{10+i}'].font = FN11

ws['B14'] = "Colour Coding"; ws['B14'].font = FB11
ws.merge_cells('B15:C15'); ws['B15'] = "Mandatory Columns"; ws['B15'].font = FW11; ws['B15'].fill = FILL_REQ
ws.merge_cells('B16:C16'); ws['B16'] = "Optional Columns"; ws['B16'].font = FW11; ws['B16'].fill = FILL_OPT
ws.merge_cells('B17:C17'); ws['B17'] = "Conditional Columns"; ws['B17'].font = FW11; ws['B17'].fill = FILL_COND

ws.merge_cells('B19:C19')
ws['B19'] = "Refund template for Export of Services with Payment of Tax"
ws['B19'].font = FB11

# ═══════════════════════════════════════
#  Header
# ═══════════════════════════════════════
ws_hd = wb.create_sheet("Header")
ws_hd.sheet_properties.tabColor = "0070C0"
ws_hd.column_dimensions['A'].width = 28
ws_hd.column_dimensions['B'].width = 30
ws_hd.column_dimensions['C'].width = 50

for ci, v in enumerate(["Field Name", "Value", "Description"]):
    c = ws_hd.cell(row=1, column=ci+1, value=v)
    c.font = FW; c.fill = FILL_REQ; c.alignment = AC; c.border = TB

ws_hd['A2'] = "GSTIN"; ws_hd['B2'] = ""; ws_hd['C2'] = "15-character GSTIN of the refund applicant"
ws_hd['A2'].font = FB10; ws_hd['B2'].font = FN10; ws_hd['C2'].font = FN10
for c in range(1, 4): ws_hd.cell(row=2, column=c).border = TB

# ═══════════════════════════════════════
#  Data
# ═══════════════════════════════════════
ws_d = wb.create_sheet("Data")
ws_d.sheet_properties.tabColor = "E67E22"

cols = [
    ("Doc Type", 14, FILL_REQ),
    ("Doc No", 18, FILL_REQ),
    ("Doc Date", 14, FILL_REQ),
    ("Doc Value", 16, FILL_REQ),
    ("Taxable Value", 16, FILL_REQ),
    ("IGST", 14, FILL_REQ),
    ("Cess", 14, FILL_OPT),
    ("BRC No", 18, FILL_COND),
    ("BRC Date", 14, FILL_COND),
    ("BRC Value", 16, FILL_COND),
    ("BRC Group ID", 14, FILL_OPT),
]

for ci, (name, w, fill) in enumerate(cols):
    col = ci + 1
    ws_d.column_dimensions[get_column_letter(col)].width = w
    cell = ws_d.cell(row=1, column=col, value=name)
    cell.font = FW; cell.fill = fill; cell.alignment = AC; cell.border = TB

sample = ["Invoice", "SVC/2024/001", "15-08-2024", 500000, 450000, 81000, "", "FIRC2024001", "20-09-2024", 450000, ""]
for ci, v in enumerate(sample):
    cell = ws_d.cell(row=2, column=ci+1, value=v)
    cell.font = FN10; cell.border = TB

dv_doc = DataValidation(type="list", formula1='"Invoice,Debit Note,Credit Note"', allow_blank=True)
ws_d.add_data_validation(dv_doc)
dv_doc.add('A2:A10001')

ws_d.freeze_panes = 'A2'

# ═══════════════════════════════════════
#  Help
# ═══════════════════════════════════════
ws_h = wb.create_sheet("Help")
ws_h.sheet_properties.tabColor = "FFC000"
ws_h.column_dimensions['A'].width = 14
ws_h.column_dimensions['B'].width = 24
ws_h.column_dimensions['C'].width = 18
ws_h.column_dimensions['D'].width = 75
ws_h.column_dimensions['E'].width = 25

r = 2

# Section 1: Header
ws_h.merge_cells(f'A{r}:E{r}')
ws_h[f'A{r}'] = "1. Sheet Name: Header"; ws_h[f'A{r}'].font = FSEC; ws_h[f'A{r}'].fill = FILL_SECTION
r += 1
ws_h.merge_cells(f'A{r}:E{r}')
ws_h[f'A{r}'] = "The following table specifies the header fields for the refund application."
ws_h[f'A{r}'].font = FN11; r += 1
for ci, h in enumerate(["Field Category", "Field Name", "Field Type", "Description", "Validation"]):
    ws_h.cell(row=r, column=ci+1, value=h).font = FHELP_H; ws_h.cell(row=r, column=ci+1).border = TB
r += 1
r = help_row(ws_h, r, "Required", "GSTIN", "Text (15)", "15-character GSTIN of the refund applicant.", "GSTIN regex; auto upper-cased and trimmed")
r += 1

# Note about return periods
ws_h.merge_cells(f'A{r}:E{r}')
ws_h[f'A{r}'] = "Note: Return Period (From/To) is NOT required for Statement 2 (v3.0). It was removed from JSON in the April 2025 update."
ws_h[f'A{r}'].font = Font(name='Calibri', size=10, italic=True); r += 2

# Section 2: Data
ws_h.merge_cells(f'A{r}:E{r}')
ws_h[f'A{r}'] = "2. Sheet Name: Data"; ws_h[f'A{r}'].font = FSEC; ws_h[f'A{r}'].fill = FILL_SECTION
r += 1
ws_h.merge_cells(f'A{r}:E{r}')
ws_h[f'A{r}'] = "The following table specifies the fields for export service invoices, tax paid, and BRC/FIRC remittance proof."
ws_h[f'A{r}'].font = FN11; r += 1
for ci, h in enumerate(["Field Category", "Field Name", "Field Type", "Description", "Validation"]):
    ws_h.cell(row=r, column=ci+1, value=h).font = FHELP_H; ws_h.cell(row=r, column=ci+1).border = TB
r += 1

r = help_row(ws_h, r, "Required", "Doc Type", "Text", "Type of document.\nValid values: Invoice, Debit Note, Credit Note", "Dropdown provided")
r = help_row(ws_h, r, "Required", "Doc No", "Text (16)", "Document number.\nMax 16 alphanumeric characters, / and - allowed.", "Cannot be 0 or special chars only")
r = help_row(ws_h, r, "Required", "Doc Date", "Text (DD-MM-YYYY)", "Date of document.\nMust be on or after 01-07-2017. Cannot be a future date.", "Valid calendar date")
r = help_row(ws_h, r, "Required", "Doc Value", "Number (15,2)", "Total value of the document (including tax).\nMust be >= 0.", "Non-negative")
r = help_row(ws_h, r, "Required", "Taxable Value", "Number (15,2)", "Taxable value of the export service.\nMust be >= 0.", "Non-negative")
r = help_row(ws_h, r, "Required", "IGST", "Number (15,2)", "Integrated Tax paid on the export service.\nMust be >= 0 and <= Taxable Value.", "Non-negative; <= Taxable Value")
r = help_row(ws_h, r, "Optional", "Cess", "Number (15,2)", "Cess amount if applicable.\nDefaults to 0 in JSON if blank.", "Non-negative if provided")
r = help_row(ws_h, r, "Conditional", "BRC No", "Text (2-30)", "Bank Realisation Certificate / Foreign Inward Remittance Certificate number.\nMandatory for Invoice/Debit Note rows that have no adjacent BRC row following.", "2-30 alphanumeric characters")
r = help_row(ws_h, r, "Conditional", "BRC Date", "Text (DD-MM-YYYY)", "Date of BRC/FIRC.\nCan be before 01-07-2017 (unlike document dates).", "Valid calendar date")
r = help_row(ws_h, r, "Conditional", "BRC Value", "Number (15,2)", "Value realised per BRC/FIRC.\nMust be >= 0 if provided.", "Non-negative")
r = help_row(ws_h, r, "Optional", "BRC Group ID", "Text", "Explicit linking identifier for grouping invoices to a single BRC/FIRC.\nUsed only with --brc-mode group option.", "Free text; used for grouping only")
r += 1

# Section 3: BRC Linking
ws_h.merge_cells(f'A{r}:E{r}')
ws_h[f'A{r}'] = "3. BRC Linking Modes"; ws_h[f'A{r}'].font = FSEC; ws_h[f'A{r}'].fill = FILL_SECTION
r += 2

brc_text = [
    "Mode 1: ADJACENT (Default)",
    "  Run: python main.py input.xlsx",
    "  BRC on a row covers ALL invoices above it (until previous BRC row).",
    "  Example:",
    "    Row 2: Invoice SVC/001 (no BRC)     ─┐",
    "    Row 3: Invoice SVC/002 (no BRC)      ├─ Covered by BRC on Row 4",
    "    Row 4: Invoice SVC/003 + BRC ₹15L   ─┘",
    "",
    "Mode 2: GROUP ID",
    "  Run: python main.py input.xlsx --brc-mode group",
    "  Use BRC Group ID column to explicitly link invoices to a BRC.",
    "  Example:",
    "    Row 2: Invoice SVC/001, Group=GRP1 (no BRC)",
    "    Row 3: Invoice SVC/002, Group=GRP1, BRC=FIRC001  → Covers both",
]
for line in brc_text:
    ws_h.merge_cells(f'A{r}:E{r}')
    ws_h[f'A{r}'] = line
    ws_h[f'A{r}'].font = Font(name='Consolas', size=10) if line.startswith("  ") else FN10
    r += 1
r += 1

# Section 4: Validation Rules
ws_h.merge_cells(f'A{r}:E{r}')
ws_h[f'A{r}'] = "4. Key Validation Rules"; ws_h[f'A{r}'].font = FSEC; ws_h[f'A{r}'].fill = FILL_SECTION
r += 2

rules = [
    "• Doc Value must be >= 0.",
    "• Taxable Value must be >= 0.",
    "• IGST must be >= 0 and <= Taxable Value.",
    "• Cess must be >= 0 if provided. Defaults to 0 in JSON if blank.",
    "• Doc Value must be >= (Taxable Value + IGST + Cess).",
    "• BRC/FIRC date CAN be before 01-07-2017 (unlike document dates).",
    "• Duplicate key = Doc Type suffix (last 4 chars) + Doc No + Financial Year.",
    "• If invoice month is Jan-Mar, FY year is decremented by 1.",
    "• Credit Notes subtract from running totals. Invoice/Debit Note add.",
]
for rule in rules:
    ws_h.merge_cells(f'A{r}:E{r}')
    ws_h[f'A{r}'] = rule; ws_h[f'A{r}'].font = FN10; r += 1
r += 1

# Section 5: Running the Tool
ws_h.merge_cells(f'A{r}:E{r}')
ws_h[f'A{r}'] = "5. Running the Tool"; ws_h[f'A{r}'].font = FSEC; ws_h[f'A{r}'].fill = FILL_SECTION
r += 2

for cmd in [
    "python main.py                         → Opens file picker",
    "python main.py input.xlsx              → Process specific file",
    "python main.py input.xlsx --pretty     → Pretty-print JSON output",
    "python main.py input.xlsx -v           → Verbose validation output",
]:
    ws_h.merge_cells(f'A{r}:E{r}')
    ws_h[f'A{r}'] = f"  {cmd}"; ws_h[f'A{r}'].font = Font(name='Consolas', size=10); r += 1

ws_h.freeze_panes = 'A3'
wb.active = wb.sheetnames.index("Overview")
wb.save(os.path.join(os.getcwd(), "GST_Refund_Stmt2_Template.xlsx"))
output = os.path.join(os.getcwd(), "GST_Refund_Stmt2_Template.xlsx")
print(f"Created: {output}")
