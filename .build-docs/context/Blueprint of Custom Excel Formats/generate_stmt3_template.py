#!/usr/bin/env python3
"""
GST Refund Offline Utility — Template Generator
Statement 3 — Exports without Payment of Tax (EXPWOP)

Generates: GST_Refund_Stmt3_Template.xlsx
Usage:     python generate_stmt3_template.py
Requires:  pip install openpyxl

The generated template follows OctaSales design language:
  - Blue headers (0070C0) = Mandatory
  - Green headers (00B050) = Optional  
  - Light blue headers (00B0F0) = Conditional
  - Calibri font throughout
  - Data starts at row 2 (no gap row)
  - Dropdowns on type/classification columns
"""
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

wb = openpyxl.Workbook()

# ─── OctaSales colour scheme ───
FILL_REQ = PatternFill('solid', fgColor='0070C0')
FILL_OPT = PatternFill('solid', fgColor='00B050')
FILL_COND = PatternFill('solid', fgColor='00B0F0')
FILL_SECTION = PatternFill('solid', fgColor='FFC000')
FILL_TITLE = PatternFill('solid', fgColor='0070C0')

FW = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
FW11 = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
FB11 = Font(name='Calibri', bold=True, size=11)
FN11 = Font(name='Calibri', size=11)
FN10 = Font(name='Calibri', size=10)
FB10 = Font(name='Calibri', bold=True, size=10)
FTITLE = Font(name='Calibri', bold=True, color='FFFFFF', size=20)
FSEC = Font(name='Calibri', bold=True, size=14)
FHELP_H = Font(name='Calibri', bold=True, size=11)
FMONO = Font(name='Consolas', size=10)

AC = Alignment(horizontal='center', vertical='center', wrap_text=True)
ALW = Alignment(horizontal='left', vertical='top', wrap_text=True)
TB = Border(left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9'))

def help_row(ws, row, cat, name, ftype, desc, val=""):
    fills = {"Required": FILL_REQ, "Optional": FILL_OPT, "Conditional": FILL_COND}
    for ci, v in enumerate([cat, name, ftype, desc, val]):
        cell = ws.cell(row=row, column=ci+1, value=v)
        cell.font = FN10; cell.border = TB; cell.alignment = ALW
        if ci == 0 and cat in fills:
            cell.fill = fills[cat]
            cell.font = Font(name='Calibri', size=10, color='FFFFFF')
    return row + 1

def help_section(ws, r, title):
    ws.merge_cells(f'A{r}:E{r}')
    ws[f'A{r}'] = title; ws[f'A{r}'].font = FSEC; ws[f'A{r}'].fill = FILL_SECTION
    return r + 1

def help_desc(ws, r, text):
    ws.merge_cells(f'A{r}:E{r}')
    ws[f'A{r}'] = text; ws[f'A{r}'].font = FN11
    return r + 1

def help_table_header(ws, r):
    for ci, h in enumerate(["Field Category", "Field Name", "Field Type", "Description", "Validation"]):
        ws.cell(row=r, column=ci+1, value=h).font = FHELP_H
        ws.cell(row=r, column=ci+1).border = TB
    return r + 1

def help_note(ws, r, text):
    ws.merge_cells(f'A{r}:E{r}')
    ws[f'A{r}'] = text; ws[f'A{r}'].font = FN10
    return r + 1

# ═══════════════════════════════════════
#  Overview
# ═══════════════════════════════════════
ws = wb.active
ws.title = "Overview"
ws.sheet_properties.tabColor = "0070C0"
ws.column_dimensions['A'].width = 3
ws.column_dimensions['B'].width = 25
ws.column_dimensions['C'].width = 65

ws.merge_cells('B2:C2')
ws['B2'] = "Refund Statement 3 Template"
ws['B2'].font = FTITLE; ws['B2'].fill = FILL_TITLE

for r, (label, val) in enumerate([
    ("Refund Type", "Exports of Goods/Services — without Payment of Tax (Accumulated ITC)"),
    ("Legal Basis", "Section 54(3) CGST Act, Rule 89(2)(b), Rule 89(2)(c), Rule 89(4)"),
    ("Refund Code / Version", "EXPWOP / 3.0"),
], start=5):
    ws[f'B{r}'] = label; ws[f'C{r}'] = val
    ws[f'B{r}'].font = FW11; ws[f'B{r}'].fill = FILL_REQ; ws[f'C{r}'].font = FN11

ws['B9'] = "Sheet Name"; ws['C9'] = "Description"
ws['B9'].font = FW11; ws['B9'].fill = FILL_REQ
ws['C9'].font = FW11; ws['C9'].fill = FILL_REQ

for i, (s, d) in enumerate([
    ("Header", "GSTIN and Return Period details"),
    ("Goods", "Export of goods — invoices, shipping/customs evidence, optional BRC/FIRC"),
    ("Services", "Export of services — invoices and BRC/FIRC remittance proof"),
    ("Help", "Field specifications, validation rules, BRC linking modes, and usage guidance"),
]):
    ws[f'B{10+i}'] = s; ws[f'C{10+i}'] = d
    ws[f'B{10+i}'].font = FW11; ws[f'B{10+i}'].fill = FILL_REQ; ws[f'C{10+i}'].font = FN11

ws['B15'] = "Colour Coding"; ws['B15'].font = FB11
ws.merge_cells('B16:C16'); ws['B16'] = "Mandatory Columns"; ws['B16'].font = FW11; ws['B16'].fill = FILL_REQ
ws.merge_cells('B17:C17'); ws['B17'] = "Optional Columns"; ws['B17'].font = FW11; ws['B17'].fill = FILL_OPT
ws.merge_cells('B18:C18'); ws['B18'] = "Conditional Columns"; ws['B18'].font = FW11; ws['B18'].fill = FILL_COND

ws.merge_cells('B20:C20')
ws['B20'] = "Refund template for Exports of Goods/Services without Payment of Tax"
ws['B20'].font = FB11

# ═══════════════════════════════════════
#  Header
# ═══════════════════════════════════════
ws_hd = wb.create_sheet("Header")
ws_hd.sheet_properties.tabColor = "0070C0"
ws_hd.column_dimensions['A'].width = 28
ws_hd.column_dimensions['B'].width = 30
ws_hd.column_dimensions['C'].width = 50

for ci, v in enumerate(["Field Name", "Value", "Description"]):
    c = ws_hd.cell(row=1, column=ci+1, value=v)
    c.font = FW; c.fill = FILL_REQ; c.alignment = AC; c.border = TB

fields = [
    ("GSTIN", "", "15-character GSTIN of the refund applicant"),
    ("Legal Name", "", "Legal name of the taxpayer (optional — for reference only)"),
    ("From Period (MM-YYYY)", "", "Starting return period e.g., 01-2025 for January 2025"),
    ("To Period (MM-YYYY)", "", "Ending return period e.g., 03-2025 for March 2025"),
]
for i, (field, val, desc) in enumerate(fields):
    r = i + 2
    ws_hd.cell(row=r, column=1, value=field).font = FB10
    ws_hd.cell(row=r, column=2, value=val).font = FN10
    ws_hd.cell(row=r, column=3, value=desc).font = FN10
    for c in range(1, 4): ws_hd.cell(row=r, column=c).border = TB

# ═══════════════════════════════════════
#  Goods
# ═══════════════════════════════════════
ws_g = wb.create_sheet("Goods")
ws_g.sheet_properties.tabColor = "27AE60"

goods_cols = [
    ("Doc Type", 12, FILL_REQ),
    ("Doc No", 16, FILL_REQ),
    ("Doc Date", 12, FILL_REQ),
    ("Doc Value", 14, FILL_REQ),
    ("Port Code", 10, FILL_REQ),
    ("SB No", 10, FILL_REQ),
    ("SB Date", 12, FILL_REQ),
    ("FOB Value", 14, FILL_REQ),
    ("EGM Ref No", 14, FILL_COND),
    ("EGM Date", 12, FILL_COND),
    ("BRC No", 16, FILL_OPT),
    ("BRC Date", 12, FILL_OPT),
    ("BRC Value", 14, FILL_OPT),
    ("BRC Group ID", 12, FILL_OPT),
]

for ci, (name, w, fill) in enumerate(goods_cols):
    col = ci + 1
    ws_g.column_dimensions[get_column_letter(col)].width = w
    cell = ws_g.cell(row=1, column=col, value=name)
    cell.font = FW; cell.fill = fill; cell.alignment = AC; cell.border = TB

# TEXT format on text-type columns — preserves leading zeros
for text_col in ['B', 'F', 'I', 'K']:  # Doc No, SB No, EGM Ref No, BRC No
    ws_g.column_dimensions[text_col].number_format = '@'

sample_g = ["Invoice", "EXP/001", "01-12-2024", 100000, "INBOM4", "1234567", "05-12-2024", 95000.5, "EGM2024001", "06-12-2024", "", "", "", ""]
for ci, v in enumerate(sample_g):
    cell = ws_g.cell(row=2, column=ci+1, value=v if v != "" else None)
    cell.font = FN10; cell.border = TB

dv_doc_g = DataValidation(type="list", formula1='"Invoice,Debit Note,Credit Note"', allow_blank=True)
ws_g.add_data_validation(dv_doc_g)
dv_doc_g.add('A2:A10001')
ws_g.freeze_panes = 'A2'

# ═══════════════════════════════════════
#  Services
# ═══════════════════════════════════════
ws_s = wb.create_sheet("Services")
ws_s.sheet_properties.tabColor = "E67E22"

svc_cols = [
    ("Doc Type", 12, FILL_REQ),
    ("Doc No", 16, FILL_REQ),
    ("Doc Date", 12, FILL_REQ),
    ("Doc Value", 14, FILL_REQ),
    ("BRC No", 16, FILL_COND),
    ("BRC Date", 12, FILL_COND),
    ("BRC Value", 14, FILL_COND),
    ("BRC Group ID", 12, FILL_OPT),
]

for ci, (name, w, fill) in enumerate(svc_cols):
    col = ci + 1
    ws_s.column_dimensions[get_column_letter(col)].width = w
    cell = ws_s.cell(row=1, column=col, value=name)
    cell.font = FW; cell.fill = fill; cell.alignment = AC; cell.border = TB

# TEXT format on text-type columns — preserves leading zeros
for text_col in ['B', 'E']:  # Doc No, BRC No
    ws_s.column_dimensions[text_col].number_format = '@'

# Sample data — showing multiple patterns
samples_s = [
    ["Invoice", "SVC/001", "01-12-2024", 500000, None, None, None, None],
    ["Invoice", "SVC/002", "15-12-2024", 750000, "FIRC2024001", "20-12-2024", 1250000, None],
    ["Invoice", "SVC/003", "20-12-2024", 300000, None, None, None, "GRP1"],
    ["Invoice", "SVC/004", "21-12-2024", 400000, "FIRC2024002", "25-12-2024", 700000, "GRP1"],
]
for ri, row_data in enumerate(samples_s):
    for ci, v in enumerate(row_data):
        cell = ws_s.cell(row=ri+2, column=ci+1, value=v)
        cell.font = FN10; cell.border = TB

dv_doc_s = DataValidation(type="list", formula1='"Invoice,Debit Note,Credit Note"', allow_blank=True)
ws_s.add_data_validation(dv_doc_s)
dv_doc_s.add('A2:A10001')
ws_s.freeze_panes = 'A2'

# ═══════════════════════════════════════
#  Help (OctaSales style)
# ═══════════════════════════════════════
ws_h = wb.create_sheet("Help")
ws_h.sheet_properties.tabColor = "FFC000"
ws_h.column_dimensions['A'].width = 14
ws_h.column_dimensions['B'].width = 24
ws_h.column_dimensions['C'].width = 18
ws_h.column_dimensions['D'].width = 75
ws_h.column_dimensions['E'].width = 25

r = 2

# ─── Section 1: Header ───
r = help_section(ws_h, r, "1. Sheet Name: Header")
r = help_desc(ws_h, r, "The following table specifies the header fields for the refund application.")
r = help_table_header(ws_h, r)
r = help_row(ws_h, r, "Required", "GSTIN", "Text (15)", "15-character GSTIN of the refund applicant.", "GSTIN regex; auto upper-cased and trimmed")
r = help_row(ws_h, r, "Optional", "Legal Name", "Text", "Legal name of the taxpayer. For reference only — not included in JSON.", "Not validated")
r = help_row(ws_h, r, "Required", "From Period", "Text (MM-YYYY)", "Starting return period.\nExample: 01-2025 for January 2025", "Must be >= 07-2017; <= current; <= To Period")
r = help_row(ws_h, r, "Required", "To Period", "Text (MM-YYYY)", "Ending return period.\nExample: 03-2025 for March 2025", "Must be >= 07-2017; <= current; >= From Period")
r += 1

# ─── Section 2: Goods ───
r = help_section(ws_h, r, "2. Sheet Name: Goods")
r = help_desc(ws_h, r, "Details of goods exported without payment of tax. One row = one export document. The G/S flag is implicit — this sheet = Goods.")
r = help_table_header(ws_h, r)
r = help_row(ws_h, r, "Required", "Doc Type", "Text", "Type of document.\nValid values: Invoice, Debit Note, Credit Note", "Dropdown provided")
r = help_row(ws_h, r, "Required", "Doc No", "Text (16)", "Document number.\nMax 16 alphanumeric characters, / and - allowed.", "Cannot be 0 or special chars only")
r = help_row(ws_h, r, "Required", "Doc Date", "Text (DD-MM-YYYY)", "Date of document.\nMust be on or after 01-07-2017. Cannot be future.", "Valid calendar date; <= To Period")
r = help_row(ws_h, r, "Required", "Doc Value", "Number (15,2)", "Total document value.\nMust be >= 0.", "Non-negative")
r = help_row(ws_h, r, "Required", "Port Code", "Text (6)", "Port code for the export.\nExactly 6 alphanumeric characters. Example: INBOM4", "Regex: ^[a-zA-Z0-9]{6}$")
r = help_row(ws_h, r, "Required", "SB No", "Text (3-7)", "Shipping Bill / Bill of Export number.\n3 to 7 digits only.", "Regex: ^[0-9]{3,7}$")
r = help_row(ws_h, r, "Required", "SB Date", "Text (DD-MM-YYYY)", "Shipping Bill date.\nCannot be future.", "Valid calendar date")
r = help_row(ws_h, r, "Required", "FOB Value", "Number (15,4)", "FOB (Free on Board) value.\nMax 15 digits + 4 decimals. Note: 4 decimals, not 2.", "Non-negative; special 4-decimal precision")
r = help_row(ws_h, r, "Conditional", "EGM Ref No", "Text (1-20)", "Export General Manifest reference number.\n1 to 20 alphanumeric characters.", "Regex: ^[a-zA-Z0-9]{1,20}$")
r = help_row(ws_h, r, "Conditional", "EGM Date", "Text (DD-MM-YYYY)", "EGM date.\nCan be before 01-07-2017.", "Valid calendar date")
r = help_row(ws_h, r, "Optional", "BRC No", "Text (2-30)", "BRC/FIRC number.\nOptional for goods exports. 2-30 alphanumeric characters.", "Regex: ^[a-zA-Z0-9]{2,30}$")
r = help_row(ws_h, r, "Optional", "BRC Date", "Text (DD-MM-YYYY)", "BRC/FIRC date.\nCan be before 01-07-2017.", "Valid calendar date")
r = help_row(ws_h, r, "Optional", "BRC Value", "Number (15,2)", "BRC/FIRC value realised.", "Non-negative if provided")
r = help_row(ws_h, r, "Optional", "BRC Group ID", "Text", "Explicit linking identifier for grouping invoices to a single BRC/FIRC.\nUsed only with --brc-mode group option.", "Free text; grouping only")
r += 1

# ─── Section 3: Services ───
r = help_section(ws_h, r, "3. Sheet Name: Services")
r = help_desc(ws_h, r, "Details of services exported without payment of tax. One row = one export document. The G/S flag is implicit — this sheet = Services.")
r = help_table_header(ws_h, r)
r = help_row(ws_h, r, "Required", "Doc Type", "Text", "Type of document.\nValid values: Invoice, Debit Note, Credit Note", "Dropdown provided")
r = help_row(ws_h, r, "Required", "Doc No", "Text (16)", "Document number.\nMax 16 alphanumeric characters, / and - allowed.", "Cannot be 0 or special chars only")
r = help_row(ws_h, r, "Required", "Doc Date", "Text (DD-MM-YYYY)", "Date of document.\nMust be on or after 01-07-2017. Cannot be future.", "Valid calendar date; <= To Period")
r = help_row(ws_h, r, "Required", "Doc Value", "Number (15,2)", "Total document value.\nMust be >= 0.", "Non-negative")
r = help_row(ws_h, r, "Conditional", "BRC No", "Text (2-30)", "BRC/FIRC number.\nMandatory for services — refund only after payment received.\nOne BRC can cover multiple invoices.", "2-30 alphanumeric characters")
r = help_row(ws_h, r, "Conditional", "BRC Date", "Text (DD-MM-YYYY)", "BRC/FIRC date.\nCan be before 01-07-2017.", "Valid calendar date")
r = help_row(ws_h, r, "Conditional", "BRC Value", "Number (15,2)", "BRC/FIRC value realised.", "Non-negative")
r = help_row(ws_h, r, "Optional", "BRC Group ID", "Text", "Explicit linking identifier for grouping invoices to a single BRC/FIRC.\nUsed only with --brc-mode group option.", "Free text; grouping only")
r += 1

# ─── Section 4: BRC Linking Modes ───
r = help_section(ws_h, r, "4. BRC Linking Modes")
r += 1
brc_lines = [
    "Mode 1: ADJACENT (Default)",
    "  Run: python main.py input.xlsx",
    "  BRC on a row covers ALL invoices above it (until previous BRC row).",
    "  Example:",
    "    Row 2: Invoice SVC/001 (no BRC)     ─┐",
    "    Row 3: Invoice SVC/002 (no BRC)      ├─ Covered by BRC on Row 4",
    "    Row 4: Invoice SVC/003 + BRC ₹15L   ─┘",
    "",
    "Mode 2: GROUP ID",
    "  Run: python main.py input.xlsx --brc-mode group",
    "  Use BRC Group ID column to explicitly link invoices to a BRC.",
    "  Example:",
    "    Row 2: Invoice SVC/001, Group=GRP1 (no BRC)",
    "    Row 3: Invoice SVC/002, Group=GRP1, BRC=FIRC001  → Covers both",
    "",
    "Applies to both Goods and Services sheets (BRC is optional for Goods, conditional for Services).",
]
for line in brc_lines:
    r = help_note(ws_h, r, line)
r += 1

# ─── Section 5: Key Validation Rules ───
r = help_section(ws_h, r, "5. Key Validation Rules")
r += 1
rules = [
    "• FOB Value allows 4 decimal places (15+4). All other amounts are 15+2.",
    "• Port Code: exactly 6 alphanumeric characters.",
    "• Shipping Bill Number: 3-7 digits only.",
    "• EGM Reference: 1-20 alphanumeric characters.",
    "• BRC/FIRC Number: 2-30 alphanumeric characters.",
    "• BRC/FIRC date and EGM date CAN be before 01-07-2017 (unlike document dates).",
    "• Document dates must be >= 01-07-2017 and <= current date and <= To Period.",
    "• Duplicate key: Doc Type suffix (last 4 chars) + Doc No + Financial Year.",
    "• If invoice month is Jan-Mar, FY year is decremented by 1.",
    "• Services: BRC/FIRC is mandatory. At least one BRC must cover each invoice.",
    "• Goods: BRC/FIRC is optional but if provided, it is included in JSON.",
    "• Goods JSON includes: sbpcode, sbnum, sbdt, fobValue, egmref, egmrefdt.",
    "• Services JSON omits all shipping/customs fields.",
    "• App auto-generates serial numbers. Sr. No. is not in the template.",
]
for rule in rules:
    r = help_note(ws_h, r, rule)
r += 1

# ─── Section 6: Running the Tool ───
r = help_section(ws_h, r, "6. Running the Tool")
r += 1
for cmd in [
    "python main.py                         → Opens file picker",
    "python main.py input.xlsx              → Process specific file",
    "python main.py input.xlsx --pretty     → Pretty-print JSON output",
    "python main.py input.xlsx -v           → Verbose validation output",
    "python main.py input.xlsx --brc-mode group  → Use Group ID linking",
]:
    ws_h.merge_cells(f'A{r}:E{r}')
    ws_h[f'A{r}'] = f"  {cmd}"; ws_h[f'A{r}'].font = FMONO; r += 1

ws_h.freeze_panes = 'A3'
wb.active = wb.sheetnames.index("Overview")
wb.save(os.path.join(os.getcwd(), "GST_Refund_Stmt3_Template.xlsx"))
output = os.path.join(os.getcwd(), "GST_Refund_Stmt3_Template.xlsx")
print(f"Created: {output}")
