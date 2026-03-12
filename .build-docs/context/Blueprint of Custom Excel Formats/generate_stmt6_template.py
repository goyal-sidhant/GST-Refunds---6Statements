#!/usr/bin/env python3
"""
GST Refund Offline Utility — Template Generator
Statement 6 — Intra/Inter Correction (INTRVC)

Generates: GST_Refund_Stmt6_Template.xlsx
Usage:     python generate_stmt6_template.py
Requires:  pip install openpyxl

The generated template follows OctaSales design language:
  - Blue headers (0070C0) = Mandatory
  - Green headers (00B050) = Optional  
  - Light blue headers (00B0F0) = Conditional
  - Calibri font throughout
  - Data starts at row 2 (no gap row)
  - Dropdowns on type/classification columns
"""
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

wb = openpyxl.Workbook()

FILL_REQ = PatternFill('solid', fgColor='0070C0')
FILL_OPT = PatternFill('solid', fgColor='00B050')
FILL_COND = PatternFill('solid', fgColor='00B0F0')
FILL_SECTION = PatternFill('solid', fgColor='FFC000')
FILL_TITLE = PatternFill('solid', fgColor='0070C0')

FW = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
FW11 = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
FB11 = Font(name='Calibri', bold=True, size=11)
FN11 = Font(name='Calibri', size=11)
FN10 = Font(name='Calibri', size=10)
FB10 = Font(name='Calibri', bold=True, size=10)
FTITLE = Font(name='Calibri', bold=True, color='FFFFFF', size=20)
FSEC = Font(name='Calibri', bold=True, size=14)
FHELP_H = Font(name='Calibri', bold=True, size=11)
FMONO = Font(name='Consolas', size=10)

AC = Alignment(horizontal='center', vertical='center', wrap_text=True)
ALW = Alignment(horizontal='left', vertical='top', wrap_text=True)
TB = Border(left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9'))

# POS master list (38 entries from govt PlaceOfSupply sheet)
POS_LIST = [
    "01-Jammu and Kashmir","02-Himachal Pradesh","03-Punjab","04-Chandigarh",
    "05-Uttarakhand","06-Haryana","07-Delhi","08-Rajasthan","09-Uttar Pradesh",
    "10-Bihar","11-Sikkim","12-Arunachal Pradesh","13-Nagaland","14-Manipur",
    "15-Mizoram","16-Tripura","17-Meghalaya","18-Assam","19-West Bengal",
    "20-Jharkhand","21-Odisha","22-Chhattisgarh","23-Madhya Pradesh",
    "24-Gujarat","25-Daman and Diu","26-Dadra and Nagaraveli","27-Maharashtra",
    "29-Karnataka","30-Goa","31-Lakshadweep","32-Kerala","33-Tamil Nadu",
    "34-Puducherry","35-Andaman and Nicobar Islands","36-Telangana",
    "37-Andhra Pradesh","38-Ladakh","97-Other Territory"
]
POS_FORMULA = '"' + ','.join(POS_LIST) + '"'

def help_row(ws, row, cat, name, ftype, desc, val=""):
    fills = {"Required": FILL_REQ, "Optional": FILL_OPT, "Conditional": FILL_COND}
    for ci, v in enumerate([cat, name, ftype, desc, val]):
        cell = ws.cell(row=row, column=ci+1, value=v)
        cell.font = FN10; cell.border = TB; cell.alignment = ALW
        if ci == 0 and cat in fills:
            cell.fill = fills[cat]
            cell.font = Font(name='Calibri', size=10, color='FFFFFF')
    return row + 1

def help_section(ws, r, title):
    ws.merge_cells(f'A{r}:E{r}')
    ws[f'A{r}'] = title; ws[f'A{r}'].font = FSEC; ws[f'A{r}'].fill = FILL_SECTION
    return r + 1

def help_desc(ws, r, text):
    ws.merge_cells(f'A{r}:E{r}')
    ws[f'A{r}'] = text; ws[f'A{r}'].font = FN11
    return r + 1

def help_table_header(ws, r):
    for ci, h in enumerate(["Field Category", "Field Name", "Field Type", "Description", "Validation"]):
        ws.cell(row=r, column=ci+1, value=h).font = FHELP_H
        ws.cell(row=r, column=ci+1).border = TB
    return r + 1

def help_note(ws, r, text):
    ws.merge_cells(f'A{r}:E{r}')
    ws[f'A{r}'] = text; ws[f'A{r}'].font = FN10
    return r + 1

# ═══════════════════════════════════════
#  Overview
# ═══════════════════════════════════════
ws = wb.active
ws.title = "Overview"
ws.sheet_properties.tabColor = "0070C0"
ws.column_dimensions['A'].width = 3
ws.column_dimensions['B'].width = 28
ws.column_dimensions['C'].width = 65

ws.merge_cells('B2:C2')
ws['B2'] = "Refund Statement 6 Template"
ws['B2'].font = FTITLE; ws['B2'].fill = FILL_TITLE

for r, (label, val) in enumerate([
    ("Refund Type", "Tax paid on Intra-State supply held to be Inter-State and vice versa (Change of POS)"),
    ("Legal Basis", "Section 77(1) and Section 19 of IGST Act, Rule 89(2)(j)"),
    ("Refund Code / Version", "INTRVC / 2.0"),
], start=5):
    ws[f'B{r}'] = label; ws[f'C{r}'] = val
    ws[f'B{r}'].font = FW11; ws[f'B{r}'].fill = FILL_REQ; ws[f'C{r}'].font = FN11

ws['B9'] = "Sheet Name"; ws['C9'] = "Description"
ws['B9'].font = FW11; ws['B9'].fill = FILL_REQ
ws['C9'].font = FW11; ws['C9'].fill = FILL_REQ

for i, (s, d) in enumerate([
    ("Header", "GSTIN, Order Number, and Order Date"),
    ("Inter to Intra", "Transactions originally taxed as Inter-State, now corrected to Intra-State"),
    ("Intra to Inter", "Transactions originally taxed as Intra-State, now corrected to Inter-State"),
    ("Help", "Field specifications, validation rules, and usage guidance"),
]):
    ws[f'B{10+i}'] = s; ws[f'C{10+i}'] = d
    ws[f'B{10+i}'].font = FW11; ws[f'B{10+i}'].fill = FILL_REQ; ws[f'C{10+i}'].font = FN11

ws['B15'] = "Colour Coding"; ws['B15'].font = FB11
ws.merge_cells('B16:C16'); ws['B16'] = "Mandatory Columns"; ws['B16'].font = FW11; ws['B16'].fill = FILL_REQ
ws.merge_cells('B17:C17'); ws['B17'] = "Optional Columns"; ws['B17'].font = FW11; ws['B17'].fill = FILL_OPT
ws.merge_cells('B18:C18'); ws['B18'] = "Conditional Columns"; ws['B18'].font = FW11; ws['B18'].fill = FILL_COND

ws.merge_cells('B20:C20')
ws['B20'] = "Refund template for tax paid under wrong Inter/Intra classification (Section 77)"
ws['B20'].font = FB11

# ═══════════════════════════════════════
#  Header
# ═══════════════════════════════════════
ws_hd = wb.create_sheet("Header")
ws_hd.sheet_properties.tabColor = "0070C0"
ws_hd.column_dimensions['A'].width = 28
ws_hd.column_dimensions['B'].width = 30
ws_hd.column_dimensions['C'].width = 50

for ci, v in enumerate(["Field Name", "Value", "Description"]):
    c = ws_hd.cell(row=1, column=ci+1, value=v)
    c.font = FW; c.fill = FILL_REQ; c.alignment = AC; c.border = TB

fields = [
    ("GSTIN", "", "15-character GSTIN of the refund applicant"),
    ("Order No", "", "Order/ruling reference number (issued under Section 77(1) / Section 19 IGST)"),
    ("Order Date", "", "Date of the order (DD-MM-YYYY)"),
]
for i, (field, val, desc) in enumerate(fields):
    r = i + 2
    ws_hd.cell(row=r, column=1, value=field).font = FB10
    ws_hd.cell(row=r, column=2, value=val).font = FN10
    ws_hd.cell(row=r, column=3, value=desc).font = FN10
    for c in range(1, 4): ws_hd.cell(row=r, column=c).border = TB

# ═══════════════════════════════════════
#  Inter to Intra sheet
# ═══════════════════════════════════════
ws_ii = wb.create_sheet("Inter to Intra")
ws_ii.sheet_properties.tabColor = "2E86C1"

ii_cols = [
    ("Recipient GSTIN", 18, FILL_COND),
    ("Recipient Name", 22, FILL_COND),
    ("Doc Type", 14, FILL_REQ),
    ("Doc No", 18, FILL_REQ),
    ("Doc Date", 14, FILL_REQ),
    ("Doc Value", 16, FILL_REQ),
    ("Taxable Value", 16, FILL_REQ),
    ("IGST Paid", 14, FILL_REQ),
    ("Earlier Cess", 14, FILL_OPT),
    ("Earlier POS", 28, FILL_REQ),
    ("Correct CGST", 14, FILL_REQ),
    ("Correct SGST", 14, FILL_REQ),
    ("Correct Cess", 14, FILL_OPT),
    ("Correct POS", 28, FILL_REQ),
]

for ci, (name, w, fill) in enumerate(ii_cols):
    col = ci + 1
    ws_ii.column_dimensions[get_column_letter(col)].width = w
    cell = ws_ii.cell(row=1, column=col, value=name)
    cell.font = FW; cell.fill = fill; cell.alignment = AC; cell.border = TB

sample_ii = ["27AABCT5678Q1Z5", "", "Invoice", "INV/2024/050", "15-08-2024", 500000, 450000,
             81000, "", "27-Maharashtra", 40500, 40500, "", "27-Maharashtra"]
for ci, v in enumerate(sample_ii):
    cell = ws_ii.cell(row=2, column=ci+1, value=v if v != "" else None)
    cell.font = FN10; cell.border = TB

# Dropdowns
dv_doc_ii = DataValidation(type="list", formula1='"Invoice,Debit Note,Credit Note"', allow_blank=True)
ws_ii.add_data_validation(dv_doc_ii)
dv_doc_ii.add('C2:C10001')

dv_pos_earlier_ii = DataValidation(type="list", formula1=POS_FORMULA, allow_blank=True)
dv_pos_earlier_ii.error = "Select a valid Place of Supply"
ws_ii.add_data_validation(dv_pos_earlier_ii)
dv_pos_earlier_ii.add('J2:J10001')

dv_pos_correct_ii = DataValidation(type="list", formula1=POS_FORMULA, allow_blank=True)
ws_ii.add_data_validation(dv_pos_correct_ii)
dv_pos_correct_ii.add('N2:N10001')

ws_ii.freeze_panes = 'A2'

# ═══════════════════════════════════════
#  Intra to Inter sheet
# ═══════════════════════════════════════
ws_it = wb.create_sheet("Intra to Inter")
ws_it.sheet_properties.tabColor = "E67E22"

it_cols = [
    ("Recipient GSTIN", 18, FILL_COND),
    ("Recipient Name", 22, FILL_COND),
    ("Doc Type", 14, FILL_REQ),
    ("Doc No", 18, FILL_REQ),
    ("Doc Date", 14, FILL_REQ),
    ("Doc Value", 16, FILL_REQ),
    ("Taxable Value", 16, FILL_REQ),
    ("CGST Paid", 14, FILL_REQ),
    ("SGST Paid", 14, FILL_REQ),
    ("Earlier Cess", 14, FILL_OPT),
    ("Earlier POS", 28, FILL_REQ),
    ("Correct IGST", 14, FILL_REQ),
    ("Correct Cess", 14, FILL_OPT),
    ("Correct POS", 28, FILL_REQ),
]

for ci, (name, w, fill) in enumerate(it_cols):
    col = ci + 1
    ws_it.column_dimensions[get_column_letter(col)].width = w
    cell = ws_it.cell(row=1, column=col, value=name)
    cell.font = FW; cell.fill = fill; cell.alignment = AC; cell.border = TB

sample_it = ["", "Ramesh Kumar", "Invoice", "INV/2024/075", "20-09-2024", 300000, 250000,
             22500, 22500, "", "19-West Bengal", 45000, "", "07-Delhi"]
for ci, v in enumerate(sample_it):
    cell = ws_it.cell(row=2, column=ci+1, value=v if v != "" else None)
    cell.font = FN10; cell.border = TB

dv_doc_it = DataValidation(type="list", formula1='"Invoice,Debit Note,Credit Note"', allow_blank=True)
ws_it.add_data_validation(dv_doc_it)
dv_doc_it.add('C2:C10001')

dv_pos_earlier_it = DataValidation(type="list", formula1=POS_FORMULA, allow_blank=True)
ws_it.add_data_validation(dv_pos_earlier_it)
dv_pos_earlier_it.add('K2:K10001')

dv_pos_correct_it = DataValidation(type="list", formula1=POS_FORMULA, allow_blank=True)
ws_it.add_data_validation(dv_pos_correct_it)
dv_pos_correct_it.add('N2:N10001')

ws_it.freeze_panes = 'A2'

# ═══════════════════════════════════════
#  Help
# ═══════════════════════════════════════
ws_h = wb.create_sheet("Help")
ws_h.sheet_properties.tabColor = "FFC000"
ws_h.column_dimensions['A'].width = 14
ws_h.column_dimensions['B'].width = 24
ws_h.column_dimensions['C'].width = 18
ws_h.column_dimensions['D'].width = 75
ws_h.column_dimensions['E'].width = 25

r = 2

# Section 1: Header
r = help_section(ws_h, r, "1. Sheet Name: Header")
r = help_desc(ws_h, r, "The following table specifies the header fields for the refund application.")
r = help_table_header(ws_h, r)
r = help_row(ws_h, r, "Required", "GSTIN", "Text (15)", "15-character GSTIN of the refund applicant.", "GSTIN regex; auto upper-cased and trimmed")
r = help_row(ws_h, r, "Required", "Order No", "Text", "Order or ruling reference number issued in pursuance of Section 77(1) of CGST Act and Section 19 of IGST Act.", "Mandatory; free text")
r = help_row(ws_h, r, "Required", "Order Date", "Text (DD-MM-YYYY)", "Date of the order/ruling.", "Valid calendar date")
r += 1

# Section 2: Inter to Intra
r = help_section(ws_h, r, "2. Sheet Name: Inter to Intra")
r = help_desc(ws_h, r, "Transactions originally taxed as INTER-STATE (IGST paid), now corrected to INTRA-STATE (CGST+SGST). One row = one document.")
r = help_table_header(ws_h, r)
r = help_row(ws_h, r, "Conditional", "Recipient GSTIN", "Text (15)", "GSTIN or UIN of the recipient.\nFill for B2B transactions. Leave blank for B2C.\nApp derives B2B/B2C from GSTIN presence.", "GSTIN regex if provided")
r = help_row(ws_h, r, "Conditional", "Recipient Name", "Text (200)", "Name of recipient.\nFill for B2C transactions (when GSTIN is blank).\nOptional for B2B.", "Required if GSTIN is blank")
r = help_row(ws_h, r, "Required", "Doc Type", "Text", "Type of document.\nValid values: Invoice, Debit Note, Credit Note", "Dropdown provided")
r = help_row(ws_h, r, "Required", "Doc No", "Text (16)", "Document number.\nMax 16 alphanumeric characters, / and - allowed.", "Cannot be 0 or special chars only")
r = help_row(ws_h, r, "Required", "Doc Date", "Text (DD-MM-YYYY)", "Date of document.\nMust be on or after 01-07-2017. Cannot be future.", "Valid calendar date")
r = help_row(ws_h, r, "Required", "Doc Value", "Number (15,2)", "Total document value.\nMust be >= 0.", "Non-negative")
r = help_row(ws_h, r, "Required", "Taxable Value", "Number (15,2)", "Taxable value of the supply.\nMust be >= 0.", "Non-negative")
r = help_row(ws_h, r, "Required", "IGST Paid", "Number (15,2)", "IGST amount that was ORIGINALLY PAID (under wrong Inter-State classification).\nThis is the tax being corrected.", "Non-negative")
r = help_row(ws_h, r, "Optional", "Earlier Cess", "Number (15,2)", "Cess paid under the earlier (wrong) classification.\nOmitted from JSON if blank.", "Non-negative if provided")
r = help_row(ws_h, r, "Required", "Earlier POS", "Text", "Place of Supply as per the EARLIER (wrong) classification.\nSelect from dropdown (38 state/UT codes).", "Dropdown provided")
r = help_row(ws_h, r, "Required", "Correct CGST", "Number (15,2)", "CGST amount under the CORRECT Intra-State classification.", "Non-negative")
r = help_row(ws_h, r, "Required", "Correct SGST", "Number (15,2)", "SGST amount under the CORRECT Intra-State classification.", "Non-negative; paired with CGST")
r = help_row(ws_h, r, "Optional", "Correct Cess", "Number (15,2)", "Cess under the correct classification.\nOmitted from JSON if blank.", "Non-negative if provided")
r = help_row(ws_h, r, "Required", "Correct POS", "Text", "Place of Supply as per the CORRECT classification.\nSelect from dropdown.", "Dropdown provided")
r += 1

# Section 3: Intra to Inter
r = help_section(ws_h, r, "3. Sheet Name: Intra to Inter")
r = help_desc(ws_h, r, "Transactions originally taxed as INTRA-STATE (CGST+SGST paid), now corrected to INTER-STATE (IGST). One row = one document.")
r = help_table_header(ws_h, r)
r = help_row(ws_h, r, "Conditional", "Recipient GSTIN", "Text (15)", "GSTIN or UIN of the recipient.\nFill for B2B transactions. Leave blank for B2C.\nApp derives B2B/B2C from GSTIN presence.", "GSTIN regex if provided")
r = help_row(ws_h, r, "Conditional", "Recipient Name", "Text (200)", "Name of recipient.\nFill for B2C transactions (when GSTIN is blank).\nOptional for B2B.", "Required if GSTIN is blank")
r = help_row(ws_h, r, "Required", "Doc Type", "Text", "Type of document.\nValid values: Invoice, Debit Note, Credit Note", "Dropdown provided")
r = help_row(ws_h, r, "Required", "Doc No", "Text (16)", "Document number.\nMax 16 alphanumeric characters, / and - allowed.", "Cannot be 0 or special chars only")
r = help_row(ws_h, r, "Required", "Doc Date", "Text (DD-MM-YYYY)", "Date of document.\nMust be on or after 01-07-2017. Cannot be future.", "Valid calendar date")
r = help_row(ws_h, r, "Required", "Doc Value", "Number (15,2)", "Total document value.\nMust be >= 0.", "Non-negative")
r = help_row(ws_h, r, "Required", "Taxable Value", "Number (15,2)", "Taxable value of the supply.\nMust be >= 0.", "Non-negative")
r = help_row(ws_h, r, "Required", "CGST Paid", "Number (15,2)", "CGST amount that was ORIGINALLY PAID (under wrong Intra-State classification).", "Non-negative")
r = help_row(ws_h, r, "Required", "SGST Paid", "Number (15,2)", "SGST amount that was ORIGINALLY PAID (under wrong Intra-State classification).", "Non-negative; paired with CGST")
r = help_row(ws_h, r, "Optional", "Earlier Cess", "Number (15,2)", "Cess paid under the earlier (wrong) classification.\nOmitted from JSON if blank.", "Non-negative if provided")
r = help_row(ws_h, r, "Required", "Earlier POS", "Text", "Place of Supply as per the EARLIER (wrong) classification.\nSelect from dropdown.", "Dropdown provided")
r = help_row(ws_h, r, "Required", "Correct IGST", "Number (15,2)", "IGST amount under the CORRECT Inter-State classification.", "Non-negative")
r = help_row(ws_h, r, "Optional", "Correct Cess", "Number (15,2)", "Cess under the correct classification.\nOmitted from JSON if blank.", "Non-negative if provided")
r = help_row(ws_h, r, "Required", "Correct POS", "Text", "Place of Supply as per the CORRECT classification.\nSelect from dropdown.", "Dropdown provided")
r += 1

# Section 4: B2B vs B2C
r = help_section(ws_h, r, "4. B2B vs B2C Logic")
r += 1
b2b_rules = [
    "• B2B/B2C is NOT a column in this template — the app derives it automatically.",
    "• If Recipient GSTIN is filled → B2B. App sets invType = 'B2B' in JSON.",
    "• If Recipient GSTIN is blank → B2C. App sets invType = 'B2C' in JSON.",
    "• For B2C: Recipient Name is mandatory (someone must be identified).",
    "• For B2B: Recipient Name is optional (GSTIN identifies the party).",
    "• At least one of GSTIN or Name must be present on every row.",
]
for rule in b2b_rules:
    r = help_note(ws_h, r, rule)
r += 1

# Section 5: Key Validation Rules
r = help_section(ws_h, r, "5. Key Validation Rules")
r += 1
rules = [
    "• Every row is a COMPARISON — earlier (wrong) tax vs correct tax. Both sides are about the SAME transaction.",
    "• Earlier and Correct classifications are ALWAYS opposite (derived from sheet name — no column needed).",
    "• Place of Supply: dropdown with 38 state/UT codes. Both Earlier POS and Correct POS are mandatory.",
    "• Earlier Cess and Correct Cess are optional. Omitted from JSON when blank (not defaulted to 0).",
    "• Doc Date must be >= 01-07-2017 and <= current date.",
    "• Duplicate detection: standard Doc Type + Doc No + FY logic.",
    "• If invoice month is Jan-Mar, FY year is decremented by 1.",
    "• No Return Periods — S06 uses Order No + Order Date instead.",
    "• This is the only statement with no export/shipping/BRC evidence and no ITC accumulation.",
]
for rule in rules:
    r = help_note(ws_h, r, rule)
r += 1

# Section 6: Running the Tool
r = help_section(ws_h, r, "6. Running the Tool")
r += 1
for cmd in [
    "python main.py                         → Opens file picker",
    "python main.py input.xlsx              → Process specific file",
    "python main.py input.xlsx --pretty     → Pretty-print JSON output",
    "python main.py input.xlsx -v           → Verbose validation output",
]:
    ws_h.merge_cells(f'A{r}:E{r}')
    ws_h[f'A{r}'] = f"  {cmd}"; ws_h[f'A{r}'].font = FMONO; r += 1

ws_h.freeze_panes = 'A3'
wb.active = wb.sheetnames.index("Overview")
wb.save(os.path.join(os.getcwd(), "GST_Refund_Stmt6_Template.xlsx"))
output = os.path.join(os.getcwd(), "GST_Refund_Stmt6_Template.xlsx")
print(f"Created: {output}")
